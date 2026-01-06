# web/server.py
from __future__ import annotations

import io
import re
import uuid
import inspect
import logging
import holidays

from calendar import monthrange
from collections import Counter
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, FastAPI, HTTPException, Query, Request, UploadFile, File, Form, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from zoneinfo import ZoneInfo
from urllib.parse import urlencode

# Projektmodule
from domain.models import Task, Reminder, Appointment, Event
from infrastructure.db_repository import DbRepository
from services.filter_service import filter_items
from utils.datetime_helpers import now_utc

from typing import Annotated
from urllib.parse import urlencode

from bootstrap import make_status_service
from services.recurrence_service import expand_item
from infrastructure.ical_mapper import to_ics
from infrastructure.ical_importer import import_ics

logging.basicConfig(level=logging.DEBUG)

# zentrale Status-Service-Instanz für das gesamte Modul
status_svc = make_status_service()

app = FastAPI()
app.mount("/static", StaticFiles(directory="web/static"), name="static")
router = APIRouter()

templates = Jinja2Templates(directory="web/templates")
DB_PATH = "taskman.db"

# ====== URL-Querystring-Helfer ======
def urlencode_qs(qp) -> str:
    """
    Wandelt Starlette QueryParams (oder dict-ähnliches) in einen URL-Querystring um.
    Unterstützt Mehrfachwerte via .multi_items(), fällt andernfalls auf items() zurück.
    """
    try:
        return urlencode(list(qp.multi_items()))
    except Exception:
        try:
            return urlencode(list(qp.items()))
        except Exception:
            return ""

def format_local_weekday_de(dt, fmt_date: str = "%a %d.%m.%Y %H:%M") -> str:
    if not dt:
        return ""
    berlin = ZoneInfo("Europe/Berlin")
    dt_local = dt.astimezone(berlin)
    wd_idx = dt_local.weekday()  # 0=Montag ... 6=Sonntag
    wd_de_full = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"][wd_idx]
    # Wenn du Kurzform willst: wd_de_short = ["Mo","Di","Mi","Do","Fr","Sa","So"][wd_idx]
    return f"{wd_de_full}"

def format_local_short_weekday_de(dt, fmt_date: str = "%a %d.%m.%Y %H:%M") -> str:
    if not dt:
        return ""
    berlin = ZoneInfo("Europe/Berlin")
    dt_local = dt.astimezone(berlin)
    wd_idx = dt_local.weekday()  # 0=Montag ... 6=Sonntag
    wd_de_short = ["Mo","Di","Mi","Do","Fr","Sa","So"][wd_idx]
    return f"{wd_de_short}"

def get_next_holidays_de_ni(start_dt_utc: datetime, count: int = 8):
    """
    Liefert die nächsten 'count' Feiertage ab start_dt_utc (inkl. heute, wenn später am Tag) für Niedersachsen.
    Rückgabe: Liste von dict-Events kompatibel zu 'Nächste Ereignisse'.
    """
    if holidays is None:
        return []

    # Deutschland, Niedersachsen
    de_ni = holidays.country_holidays('DE', subdiv='NI', years=range(start_dt_utc.year, start_dt_utc.year + 2))
    out = []
    # Lauf durch nächsten 400 Tage (robust, ohne teure Suche)
    day = start_dt_utc.date()
    end_day = (start_dt_utc + timedelta(days=400)).date()
    while day <= end_day and len(out) < count:
        if day in de_ni:
            name = de_ni.get(day)
            # Ganztägig -> start=end auf 00:00 UTC
            start = datetime(day.year, day.month, day.day, 0, 0, tzinfo=timezone.utc)
            end   = start + timedelta(days=1)
            out.append({
                "id": f"holiday-{day.isoformat()}",
                "type": "event",
                "status": "EVENT_SCHEDULED",
                "name": f"{name} (Feiertag)",
                "start_utc": start,
                "end_utc": end,
                "is_holiday": True,
                "priority": 0,
                "tags": ["Feiertag", "DE", "NI"],
            })
        day += timedelta(days=1)
    return out

# ====== HTMX-Helfer ======
def is_htmx(request: Request) -> bool:
    return request.headers.get("HX-Request", "false").lower() == "true"

def hx_redirect(url: str) -> Response:
    resp = Response(status_code=204)
    resp.headers["HX-Redirect"] = url
    return resp

def hx_refresh() -> Response:
    resp = Response(status_code=204)
    resp.headers["HX-Refresh"] = "true"
    return resp
        
# ====== Jinja-Filter: Lokalzeit & deutsche Wochentage ======
def format_local(dt: Optional[datetime], fmt: str = "%d.%m.%Y %H:%M") -> str:
    if not dt:
        return ""
    try:
        return dt.astimezone(ZoneInfo("Europe/Berlin")).strftime(fmt)
    except Exception:
        return ""

def _de_weekday_map(s: str) -> str:
    # Kurzformen EN->DE
    return (s.replace("Mon", "Mo").replace("Tue", "Di").replace("Wed", "Mi")
             .replace("Thu", "Do").replace("Fri", "Fr").replace("Sat", "Sa").replace("Sun", "So"))

templates.env.filters["urlencode_qs"] = urlencode_qs
templates.env.filters["format_local"] = format_local
templates.env.filters["format_local_weekday_de"] = format_local_weekday_de
templates.env.filters["format_local_short_weekday_de"] = format_local_short_weekday_de

def has_yearly_semantics(it) -> bool:
    rec = getattr(it, "recurrence", None)
    recstr = getattr(rec, "rrulestring", None) if rec else None
    if recstr and "FREQ=YEARLY" in recstr.upper():
        return True
    tags = getattr(it, "tags", None) or []
    if is_birthday(it):
        return True
    return False

def compute_next_yearly_from(it, now: Optional[datetime] = None):
    now = now or datetime.now(timezone.utc)
    rec = getattr(it, "recurrence", None)
    recstr = getattr(rec, "rrulestring", None) if rec else None
    base = None
    if recstr:
        m = re.search(r"DTSTART:(\d{8}T\d{6}Z)", recstr)
        if m:
            try:
                base = datetime.strptime(m.group(1), "%Y%m%dT%H%M%SZ").replace(tzinfo=timezone.utc)
            except Exception:
                base = None
    if base is None:
        base = getattr(it, "start_utc", None)
    if base is None:
        return None
    # duration
    dur = timedelta(days=1)
    s0 = getattr(it, "start_utc", None)
    e0 = getattr(it, "end_utc", None)
    if s0 and e0:
        try:
            dur = max(e0 - s0, timedelta(hours=1))
        except Exception:
            pass
    y = now.year
    m = base.month
    d = base.day
    hh, mm, ss = base.hour, base.minute, base.second
    # 29.02 clamp
    try:
        cand = datetime(y, m, d, hh, mm, ss, tzinfo=timezone.utc)
    except ValueError:
        last_day = monthrange(y, m)[1]
        cand = datetime(y, m, min(d, last_day), hh, mm, ss, tzinfo=timezone.utc)
    if cand <= now:
        y += 1
        try:
            cand = datetime(y, m, d, hh, mm, ss, tzinfo=timezone.utc)
        except ValueError:
            last_day = monthrange(y, m)[1]
            cand = datetime(y, m, min(d, last_day), hh, mm, ss, tzinfo=timezone.utc)
    return cand, cand + dur

def next_or_display_occurrence(it, now: Optional[datetime] = None, require_future: bool = True):
    now = now or datetime.now(timezone.utc)
    t = getattr(it, "type", None)
    # tasks/reminders: show current until terminal; next only after completion
    if t in ("task", "reminder"):
        due_or_rem = getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None)
        return due_or_rem, None
    if t == "appointment":
        s = getattr(it, "start_utc", None)
        e = getattr(it, "end_utc", None)
        if status_svc.is_terminal(it.status):
            rec = getattr(it, "recurrence", None)
            recstr = getattr(rec, "rrulestring", None) if rec else None
            if recstr:
                nxt = _next_occurrence_from_rrule(recstr, it, now)  # existing helper
                if nxt:
                    return nxt
        return s, e
    if t == "event":
        rec = getattr(it, "recurrence", None)
        recstr = getattr(rec, "rrulestring", None) if rec else None
        if recstr:
            nxt = _next_occurrence_from_rrule(recstr, it, now)  # existing helper
            if nxt:
                return nxt
            if "FREQ=YEARLY" in (recstr or "").upper():
                y = compute_next_yearly_from(it, now)
                if y:
                    return y
        if has_yearly_semantics(it):
            y = compute_next_yearly_from(it, now)
            if y:
                return y
        return getattr(it, "start_utc", None), getattr(it, "end_utc", None)
    return getattr(it, "start_utc", None), getattr(it, "end_utc", None)

# ====== DI ======
def get_repo():
    repo = DbRepository("taskman.db")
    try:
        yield repo
    finally:
        try:
            repo.conn.close()
        except Exception:
            pass

def get_status():
    svc = make_status_service()
    return svc

# ====== Status-Helfer ======
def _status_options_for(status, item_type: str):
    if hasattr(status, "get_options_for"):
        defs = status.get_options_for(item_type)
        out = []
        for sd in defs:
            key = getattr(sd, "key", None) if sd is not None else None
            label = getattr(sd, "display_name", None) if sd is not None else None
            if key is None or label is None:
                try:
                    key = sd[0]; label = sd[1]  # type: ignore[index]
                except Exception:
                    continue
            out.append((key, label))
        return out
    if hasattr(status, "options_for"):
        defs = status.options_for(item_type)
        out = []
        for sd in defs:
            key = getattr(sd, "key", None)
            label = getattr(sd, "display_name", None)
            if key is None or label is None:
                if isinstance(sd, tuple) and len(sd) >= 2:
                    key, label = sd[0], sd[1]
                else:
                    continue
            out.append((key, label))
        return out
    return []

def _status_colors_for(status, item_type: str) -> dict:
    colors = {}
    if hasattr(status, "get_options_for"):
        for sd in status.get_options_for(item_type):
            key = getattr(sd, "key", None)
            col = getattr(sd, "color_light", None)
            if key and col:
                colors[key] = col
    return colors

def _status_display(status, key: str) -> str:
    if hasattr(status, "reverse_format"):
        return status.reverse_format(key)
    if hasattr(status, "display_name"):
        return status.display_name(key)
    return key

# ====== Zeit-/RRULE-Helfer ======
def _parse_local_dt(s: str) -> Optional[datetime]:
    try:
        dt_local = datetime.strptime(s.strip(), "%d.%m.%Y %H:%M").replace(tzinfo=ZoneInfo("Europe/Berlin"))
        return dt_local.astimezone(ZoneInfo("UTC"))
    except Exception:
        return None

def _byday_de_to_en(rr: str) -> str:
    if not rr: return rr
    return (rr.replace("BYDAY=DI", "BYDAY=TU")
              .replace("BYDAY=MI", "BYDAY=WE")
              .replace("BYDAY=DO", "BYDAY=TH")
              .replace("BYDAY=SO", "BYDAY=SU"))

def _normalize_rrule_input(dtstart_local: str, rrule_line: str, exdates_local: str):
    dtstart_utc = _parse_local_dt(dtstart_local) if (dtstart_local or "").strip() else None
    rrule_line = _byday_de_to_en((rrule_line or "").strip())
    if not dtstart_utc and not rrule_line and not (exdates_local or "").strip():
        return None, None
    rrule_string = None
    if rrule_line:
        if dtstart_utc:
            rrule_string = f"DTSTART:{dtstart_utc.strftime('%Y%m%dT%H%M%SZ')}\nRRULE:{rrule_line}"
        else:
            rrule_string = f"RRULE:{rrule_line}"
    exdates_utc = []
    if (exdates_local or "").strip():
        for part in exdates_local.split(","):
            d = _parse_local_dt(part.strip())
            if d:
                exdates_utc.append(d)
    return rrule_string, tuple(exdates_utc) if exdates_utc else None

def _build_recurrence(rrule_string: Optional[str], exdates_utc: Optional[tuple]):
    from domain.models import Recurrence
    if not rrule_string and not exdates_utc:
        return None
    return Recurrence(rrule_string=rrule_string or "", exdates_utc=exdates_utc or ())

def _validate_edit_input(it, status, name, status_key, due, start_local, end_local, dtstart_local, rrule_line, exdates_local) -> list[str]:
    msgs = []

    # Name prüfen
    if not (name or "").strip():
        msgs.append("Name darf nicht leer sein.")

    # Erlaubte Keys je Typ aus zentralem Service
    opts = status_svc.get_options_for(getattr(it, "type", None))
    allowed_keys = {sd.key for sd in opts}  # Set für schnellen Lookup

    # Status typbewusst normalisieren (Altformen + Präfix-Garantie)
    normalized = status_svc.normalize_input(status_key or "", getattr(it, "type", None)) if status_key else None

    # Allowed-Prüfung
    if allowed_keys and normalized and normalized not in allowed_keys:
        msgs.append("Ungültiger Status für diesen Typ.")

    # Übergangsregel validieren (nur wenn ein neuer Status angegeben ist)
    if normalized:
        is_recurring = bool(getattr(it, "recurrence", None) or getattr(it, "rrule_string", None))
        ok, reason = status_svc.validate_transition(getattr(it, "status", None), normalized, is_recurring=is_recurring)

    # Datumshilfsprüfer
    def _try_dt(label: str, val: Optional[str]):
        if (val or "").strip() and _parse_local_dt(val) is None:
            msgs.append(f"Ungültiges Datum '{label}': {val}")

    # Feldprüfungen je Typ
    if it.type == "task":
        _try_dt("Fällig", due)
    elif it.type in ("appointment", "event"):
        _try_dt("Start", start_local)
        _try_dt("Ende", end_local)
    elif it.type == "reminder":
        _try_dt("Erinnerung", due)

    # RRule-Form prüfen
    if (rrule_line or "").strip():
        if not rrule_line.strip().upper().startswith("FREQ="):
            msgs.append("Wiederholung muss mit FREQ=... beginnen (z. B. FREQ=DAILY, FREQ=WEEKLY).")

    _try_dt("DTSTART", dtstart_local)

    # EXDATE-Liste prüfen
    if (exdates_local or "").strip():
        for p in exdates_local.split(","):
            _try_dt("EXDATE", (p or "").strip())

    return msgs

def build_status_choices(type_status_options: Dict[str, List[Tuple[str, str]]]) -> List[Tuple[str, str]]:
    seen = {}
    for _, opts in (type_status_options or {}).items():
        for key, label in (opts or []):
            if key not in seen:
                seen[key] = label
    order = {"open": 0, "in_progress": 1, "planned": 2, "active": 3, "done": 90, "canceled": 91}
    return sorted(seen.items(), key=lambda kv: (order.get(kv[0].lower(), 50), kv[1].lower()))

_DT_FMT = "%Y%m%dT%H%M%SZ"

def _parse_rrule(rec_str: str) -> Optional[dict]:
    if not rec_str:
        return None
    m_dt = re.search(r"DTSTART:(\d{8}T\d{6}Z)", rec_str)
    m_rr = re.search(r"RRULE:([^\r\n]+)", rec_str)
    if not (m_dt and m_rr):
        return None
    try:
        dtstart = datetime.strptime(m_dt.group(1), _DT_FMT).replace(tzinfo=timezone.utc)
    except Exception:
        return None

    parts = {}
    for kv in m_rr.group(1).split(";"):
        if "=" in kv:
            k, v = kv.split("=", 1)
            parts[k.upper()] = v

    def _i(k, d=None):
        try:
            return int(parts[k]) if k in parts else d
        except Exception:
            return d

    def _dt(k):
        try:
            return datetime.strptime(parts[k], _DT_FMT).replace(tzinfo=timezone.utc) if k in parts else None
        except Exception:
            return None

    return {
        "dtstart": dtstart,
        "freq": (parts.get("FREQ") or "").upper(),
        "interval": max(1, _i("INTERVAL", 1) or 1),
        "count": _i("COUNT", None),
        "until": _dt("UNTIL"),
        "bymonth": _i("BYMONTH", None),
        "bymonthday": _i("BYMONTHDAY", None),
        # Erweiterbar: BYDAY, BYSETPOS etc.
    }

def _duration_for(it) -> timedelta:
    s = getattr(it, "start_utc", None)
    e = getattr(it, "end_utc", None)
    if s and e:
        try:
            return max(e - s, timedelta(hours=1))
        except Exception:
            pass
    return timedelta(days=1)  # All-day Default

def _clamp_month_day(year:int, month:int, day:int) -> int:
    return min(day, monthrange(year, month)[1])

def _norm_year_date(year:int, month:int, day:int, hh:int, mm:int, ss:int) -> datetime:
    dmax = monthrange(year, month)[1]
    d = min(day, dmax)
    return datetime(year, month, d, hh, mm, ss, tzinfo=timezone.utc)

def _next_occurrence_from_rrule(rec_str: str, it, now: datetime):
    p = _parse_rrule(rec_str)
    if not p:
        return None
    dtstart = p["dtstart"]; freq = p["freq"]; interval = p["interval"]
    count = p["count"]; until = p["until"]; bym = p["bymonth"]; byd = p["bymonthday"]
    dur = _duration_for(it)

    if freq == "YEARLY":
        m = bym if bym else dtstart.month
        d = byd if byd else dtstart.day
        hh, mm, ss = dtstart.hour, dtstart.minute, dtstart.second
        y0 = dtstart.year
        y = now.year
        cand = _norm_year_date(y, m, d, hh, mm, ss)
        if cand < now:
            y += 1
            cand = _norm_year_date(y, m, d, hh, mm, ss)
        mod = (y - y0) % interval
        if mod != 0:
            y += (interval - mod)
            cand = _norm_year_date(y, m, d, hh, mm, ss)
        if count is not None:
            idx = 1 + (y - y0)//interval
            if idx > count:
                return None
        if until is not None and cand > until:
            return None
        return (cand, cand + dur)

    # Einfache Implementierungen für andere Frequenzen (optional erweitern)
    if freq == "DAILY":
        delta = now - dtstart
        days = max(0, delta.days)
        steps = (days // interval) * interval
        occ = dtstart + timedelta(days=steps)
        while occ < now:
            occ += timedelta(days=interval)
            steps += interval
            if count is not None and (1 + steps // interval) > count:
                return None
            if until is not None and occ > until:
                return None
        return (occ, occ + dur)

    if freq == "WEEKLY":
        delta_days = (now - dtstart).days
        weeks = max(0, delta_days // 7)
        steps = (weeks // interval) * interval
        occ = dtstart + timedelta(weeks=steps)
        while occ < now:
            occ += timedelta(weeks=interval)
            steps += interval
            if count is not None and (1 + steps // interval) > count:
                return None
            if until is not None and occ > until:
                return None
        return (occ, occ + dur)

    if freq == "MONTHLY":
        # grob: von dtstart in Monatsintervallen vorspringen
        total = (now.year - dtstart.year) * 12 + (now.month - dtstart.month)
        steps = max(0, (total // interval))
        # Zielmonat:
        y = dtstart.year + ((dtstart.month - 1 + steps * interval) // 12)
        m = (dtstart.month - 1 + steps * interval) % 12 + 1
        bymonthday = p["bymonthday"]
        d = bymonthday if bymonthday else dtstart.day
        hh, mm, ss = dtstart.hour, dtstart.minute, dtstart.second
        d = min(d, monthrange(y, m)[1])
        occ = datetime(y, m, d, hh, mm, ss, tzinfo=timezone.utc)
        while occ < now:
            # nächster Block
            steps += 1
            y = dtstart.year + ((dtstart.month - 1 + steps * interval) // 12)
            m = (dtstart.month - 1 + steps * interval) % 12 + 1
            d = bymonthday if bymonthday else dtstart.day
            d = min(d, monthrange(y, m)[1])
            occ = datetime(y, m, d, hh, mm, ss, tzinfo=timezone.utc)
            if count is not None and steps + 1 > count:
                return None
            if until is not None and occ > until:
                return None
        return (occ, occ + dur)

    # Nicht unterstützt -> None
    return None

def _aware(dt: datetime | None) -> datetime | None:
    """Ensure tz-aware UTC datetime; keep existing tz, attach UTC if naive."""
    if dt is None:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


# ====== Index mit Filter, berechneten Anzeige-Zeiten und korrekter Sortierung ======
@app.get("/", response_class=HTMLResponse)
def index(
    request: Request,
    q: Optional[str] = None,
    types: Optional[str] = None,
    status_keys: Optional[str] = None,   # CSV optional
    status: Optional[str] = None,        # EINZELNER Status aus dem Dropdown
    show_private: int = 0,
    include_past: int = 0,
    tags: Optional[str] = None,
    sort: Optional[str] = None,
    dir: str = "asc",
    range: Optional[str] = None,         # range Parameter explizit aufgenommen
    repo: DbRepository = Depends(get_repo),
):
    # 1. Daten laden
    items = repo.list_all()
    now_dt = now_utc()

    # ====== PIPELINE SCHRITT 1: Status-Mutation & Autokorrektur ======
    # (Dieser Block muss ganz am Anfang stehen, damit Filter auf korrekten Daten arbeiten)
    changed = False
    updated = []

    for it in list(items):
        it_type = getattr(it, "type", None)
        if not it_type:
            continue  # Typ unbekannt -> nicht anfassen

        opts = status_svc.get_options_for(it_type)
        allowed_keys = [sd.key for sd in opts] if opts else []

        cur = getattr(it, "status", "") or ""
        norm = status_svc.normalize_input(cur, it_type) if cur else ""

        def _prefix_ok(t, key):
            return (t == "task" and key.startswith("TASK_")) or \
                (t == "appointment" and key.startswith("APPOINTMENT_")) or \
                (t == "event" and key.startswith("EVENT_")) or \
                (t == "reminder" and (key.startswith("REMINDER_") or key == ""))

        if _prefix_ok(it_type, cur) and (cur in allowed_keys):
            norm = cur
        else:
            if norm in allowed_keys:
                pass
            else:
                if allowed_keys:
                    default_key = next((sd.key for sd in opts if not sd.is_terminal), allowed_keys[0])
                    norm = default_key
                else:
                    norm = cur # keine Optionen bekannt -> nichts tun

        # Auto-Finalisierung
        if it_type in ("appointment", "event"):
            payload = {}
            end_val = getattr(it, "end_utc", None) or getattr(it, "end_dt", None) or getattr(it, "end_time", None) or getattr(it, "until", None)
            start_val = getattr(it, "start_utc", None)
            
            if end_val: payload["end"] = end_val
            if start_val: payload["start"] = start_val

            if payload:
                suggested = status_svc.auto_adjust_appointment_status(payload, now=now_dt)
                
                # Geburtstags-Guard
                if it_type == "event" and suggested in {"EVENT_DONE", "EVENT_CANCELLED"}:
                    if is_birthday(it):
                        suggested = None

                if suggested and suggested in allowed_keys:
                    norm = suggested

        if norm and norm != cur:
            it2 = it.__class__(**{**it.__dict__, "status": norm})
            repo.upsert(it2)
            updated.append(it2)
            changed = True

    if changed:
        repo.conn.commit()
        idmap = {it.id: it for it in items}
        for it2 in updated:
            idmap[it2.id] = it2
        items = list(idmap.values())

    # ====== PIPELINE SCHRITT 2: Statische Filterung (DB-Attribute) ======
    
    # Parameter vorbereiten
    types_list = types.split(",") if types else None
    status_list = [s.strip() for s in status_keys.split(",")] if status_keys else None
    sel_status = (status or "").strip()
    if sel_status:
        status_list = [sel_status]

    tags_multi = [t.strip() for t in request.query_params.getlist("tags") if t.strip()]
    csv_raw = request.query_params.get("tags")
    tags_csv = [t.strip() for t in csv_raw.split(",")] if csv_raw else []
    tags_list = list(dict.fromkeys([*tags_multi, *tags_csv])) or None

    prio = request.query_params.get("prio")
    min_prio = int(prio) if prio not in (None, "") else None

    # Statisch filtern (alles was nicht Zeit/Logik betrifft)
    # Nutzt den importierten filter_items service
    try:
        items = filter_items(
            items=items,
            text=q,
            types=types_list,
            status_keys=status_list,
            include_private=bool(int(show_private or 0)), 
            tags=tags_list, 
            min_priority=min_prio,
        )
    except Exception as e:
        print(f"Filter error fallback: {e}")
        if q:
            q_norm = q.strip().lower()
            items = [it for it in items if q_norm in (getattr(it, "name", "") or "").lower()]


    # ====== PIPELINE SCHRITT 3: Zeitfenster & Occurrences Berechnen ======
    
    rng = (range or request.query_params.get("range") or "").strip().lower()
    local_tz = ZoneInfo("Europe/Berlin") # TODO: User preference
    
    # 3a. Zeitfenster definieren
    win_start, win_end = None, None
    
    if rng == "heute":
        now_local = datetime.now(local_tz)
        day_start = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
        win_start = day_start.astimezone(timezone.utc)
        win_end = (day_start + timedelta(days=1)).astimezone(timezone.utc)
        
    elif rng == "woche":
        now_local = datetime.now(local_tz)
        today_start = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
        days_until_sunday = (7 - today_start.isoweekday()) % 7
        if days_until_sunday == 0: days_until_sunday = 0 # Heute ist Sonntag
        win_start = today_start.astimezone(timezone.utc)
        win_end = (today_start + timedelta(days=days_until_sunday + 1)).astimezone(timezone.utc)
        
    elif rng == "naechstewoche":
        now_local = datetime.now(local_tz)
        today_start = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
        days_until_monday = (8 - today_start.isoweekday()) % 7
        if days_until_monday == 0: days_until_monday = 7
        next_monday = today_start + timedelta(days=days_until_monday)
        win_start = next_monday.astimezone(timezone.utc)
        win_end = (next_monday + timedelta(days=7)).astimezone(timezone.utc)

    # Hilfsfunktionen für Fenster-Logik (lokal definiert für Zugriff auf Scope)
    def filter_occurrences(occs, ws, we):
        """Filtert Occurrences nach Fenster-Überlappung"""
        filtered = []
        for occ in occs:
            if occ.item_type in ('task', 'reminder'):
                # Tasks/Reminders: due_utc muss im Fenster liegen
                # Achtung: occ.due_utc oder occ.reminder_utc nutzen
                dt = getattr(occ, 'due_utc', None) or getattr(occ, 'reminder_utc', None)
                if in_window(dt, ws, we):
                    filtered.append(occ)
            else:
                # Appointments/Events: muss Fenster überlappen
                if overlaps_window(occ.start_utc, occ.end_utc, ws, we):
                    filtered.append(occ)
        return filtered

    def expand_window_safe(it, ws, we):
        """Expandiert Item im Zeitfenster mit Überlappungsfilter"""
        # expand_item muss importiert sein (aus recurrence_service)
        raw_occs = expand_item(it, ws, we) or []
        
        # Markiere alle Occurrences dieses Items als Geburtstag (falls zutreffend)
        if is_birthday(it):
            for occ in raw_occs:
                # Hack für Frozen Dataclass/Objekt, falls nötig
                try:
                    object.__setattr__(occ, 'is_birthday', True)
                except:
                    pass 
        
        return filter_occurrences(raw_occs, ws, we)

    def in_window(dt, s, e):
        return (dt is not None) and (s <= dt < e)

    def overlaps_window(s, e, ws, we):
        if s is None or e is None: return False
        return (s < we) and (e > ws)

    def _utc_aware(dt):
        if not dt: return None
        return dt if dt.tzinfo is not None else dt.replace(tzinfo=timezone.utc)
    
    def _aware(dt): return _utc_aware(dt) # Alias

    # 3b. Rows aufbauen (Berechnung & Zeit-Filterung)
    rows = []
    
    show_past_bool = bool(int(include_past or 0))
    
    for it in items:
        occs = []
        disp_start = None
        disp_end = None
        keep_item = False

        # Fall A: Expliziter Zeitraum gewählt (Heute/Woche...)
        if win_start and win_end:
            # Nutze expand_window_safe (inkl. Birthday-Fix)
            # Hinweis: expand_window_safe muss im Scope verfügbar sein (s.u. oder importiert)
            raw_occs = expand_window_safe(it, win_start, win_end)
            if raw_occs:
                keep_item = True
                occs = raw_occs
                # Display Time vom ERSTEN Treffer im Fenster
                first = occs[0]
                if getattr(it, 'type', '') in ('task', 'reminder'):
                    disp_start = getattr(first, 'due_utc', None) or getattr(first, 'reminder_utc', None)
                else:
                    disp_start = getattr(first, 'start_utc', None)
                    disp_end = getattr(first, 'end_utc', None)

        # Fall B: Kein Zeitraum (Default View / "Alles")
        else:
            t = getattr(it, "type", "") or ""

            # _expand_next kann bei nicht-wiederkehrenden Items auch Vergangenes zurückgeben
            occs_raw = _expand_next(it, start_dt=now_dt, max_count=3)

            # "Echte" zukünftige/aktive Occurrences ableiten (für include_past=0 Entscheidung)
            real_future_occs = []
            for o in occs_raw:
                if t in ("appointment", "event"):
                    o_start = _utc_aware(getattr(o, "start_utc", None))
                    o_end = _utc_aware(getattr(o, "end_utc", None))

                    # Relevanz: noch nicht vorbei (Ende >= jetzt) oder (wenn kein Ende) Start >= jetzt
                    if o_end is not None:
                        if o_end >= now_dt:
                            real_future_occs.append(o)
                    elif o_start is not None:
                        if o_start >= now_dt:
                            real_future_occs.append(o)
                    else:
                        # Keine Zeitinformationen -> als "relevant" behandeln (defensiv)
                        real_future_occs.append(o)

                elif t in ("task", "reminder"):
                    o_dt = _utc_aware(getattr(o, "due_utc", None) or getattr(o, "reminder_utc", None))
                    if o_dt is None or o_dt >= now_dt:
                        real_future_occs.append(o)

                else:
                    # Unbekannter Typ: defensiv behalten
                    real_future_occs.append(o)

            has_future_occs = len(real_future_occs) > 0

            # Für Anzeige/Syntax: bei Terminen/Remindern nur die "echten" künftigen Occurrences nutzen,
            # Tasks dürfen überfällig sein (werden trotzdem angezeigt, wenn nicht terminal)
            if t in ("appointment", "event", "reminder"):
                occs = real_future_occs
            else:
                occs = occs_raw

            base_due = _utc_aware(getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None))
            base_start = _utc_aware(getattr(it, "start_utc", None))
            base_end = _utc_aware(getattr(it, "end_utc", None))

            st = getattr(it, "status", "") or ""
            is_terminal = status_svc.is_terminal(st)

            if show_past_bool:
                keep_item = True
            else:
                if t == "task":
                    is_recurring = bool(getattr(it, "recurrence", None) or getattr(it, "rrule_string", None))
                    if is_recurring:
                        keep_item = True
                    else:
                        keep_item = not is_terminal  # nicht-wiederkehrende erledigte Tasks ausblenden

                elif t == "reminder":
                    if is_terminal:
                        keep_item = False
                    else:
                        reminder_dt = _utc_aware(getattr(it, "reminder_utc", None))
                        if reminder_dt is None:
                            keep_item = True
                        elif reminder_dt >= now_dt:
                            keep_item = True
                        else:
                            keep_item = has_future_occs

                elif t in ("appointment", "event"):
                    if is_terminal:
                        keep_item = False
                    else:
                        end_dt = base_end
                        start_dt = base_start
                        if end_dt is not None:
                            keep_item = (end_dt >= now_dt) or has_future_occs
                        elif start_dt is not None:
                            keep_item = (start_dt >= now_dt) or has_future_occs
                        else:
                            # Zeitloser Termin/Event -> anzeigen
                            keep_item = True

                else:
                    keep_item = True

            # Display Werte (Priorität: Occurrences -> Basiswerte)
            if occs:
                first = occs[0]
                if t in ("task", "reminder"):
                    disp_start = getattr(first, "due_utc", None) or getattr(first, "reminder_utc", None)
                    disp_end = None
                else:
                    disp_start = getattr(first, "start_utc", None)
                    disp_end = getattr(first, "end_utc", None)
            else:
                if t in ("task", "reminder"):
                    disp_start = base_due
                    disp_end = None
                else:
                    disp_start = base_start
                    disp_end = base_end

        # Finale Entscheidung für dieses Item
        if keep_item:
            # Fallback falls disp_start immer noch None (z.B. Task ohne alles)
            # Für Sortierung wichtig
            rows.append((it, occs, _aware(disp_start), _aware(disp_end)))


    # ====== PIPELINE SCHRITT 4: Sortierung (Auf BERECHNETEN Daten) ======
    key = (sort or "").strip()
    reverse = (dir or "asc").lower() == "desc"
    key_norm = key.lower()

    # Default Sortierung wenn nichts gewählt
    if not key_norm:
        key_norm = "start_faellig" 

    try:
        if key_norm == "type":
            rows.sort(key=lambda r: (getattr(r[0], "type", "") or "").lower(), reverse=reverse)
        
        elif key_norm == "name":
            rows.sort(key=lambda r: (getattr(r[0], "name", "") or "").lower(), reverse=reverse)
        
        elif key_norm == "status":
            rows.sort(key=lambda r: (getattr(r[0], "status", "") or "").lower(), reverse=reverse)
        
        elif key_norm == "priority":
            def rk_prio(r):
                it, _, ds, _ = r
                p = getattr(it, "priority", None)
                # None nach hinten (999), Hohe Prio (5) vor niedriger (1) -> bei asc umgekehrt? 
                # Üblich: 5=Hoch, 1=Tief. Sort DESC für Wichtiges oben.
                p_val = 999 if p is None else p
                # Sekundär: Datum
                d_val = ds or datetime.max.replace(tzinfo=timezone.utc)
                return (p_val, d_val)
            rows.sort(key=rk_prio, reverse=reverse)

        elif key_norm == "tags":
            rows.sort(key=lambda r: len(getattr(r[0], "tags", []) or []), reverse=reverse)

        elif key_norm == "changed":
            def rk_changed(r):
                it, _, _, _ = r
                ch = _utc_aware(getattr(it, "last_modified_utc", None))
                return ch or datetime.min.replace(tzinfo=timezone.utc)
            rows.sort(key=rk_changed, reverse=reverse)

        elif key_norm.startswith("start") or key_norm in ("due","fällig","faellig","start/fällig","start_faellig"):
            def rk_date(r):
                it, _, ds, _ = r
                # Primär: Das berechnete Display-Start Datum
                # Items ohne Datum (ds=None) sollen ans Ende (bei ASC)
                has_date = 0 if ds is not None else 1
                dt_val = ds or datetime.max.replace(tzinfo=timezone.utc)
                
                # Sekundär: Name für Stabilität
                name_val = (getattr(it, "name", "") or "").lower()
                
                return (has_date, dt_val, name_val)
            
            rows.sort(key=rk_date, reverse=reverse)
        
        else:
            # Fallback Name
            rows.sort(key=lambda r: (getattr(r[0], "name", "") or "").lower(), reverse=reverse)

    except Exception as ex:
        print(f"Sort Error ({key_norm}): {ex}")
        rows.sort(key=lambda r: (getattr(r[0], "name", "") or "").lower())


    # ====== PIPELINE SCHRITT 5: UI-Helper laden & Render ======
    
    # Status Optionen laden (für UI)
    TYPES = ("task", "reminder", "appointment", "event")
    _raw = {t: status_svc.get_options_for(t) for t in TYPES}
    
    type_status_options = {t: [(sd.key, sd.display_name) for sd in defs] for t, defs in _raw.items()}
    
    status_choices = [(sd.key, sd.display_name) for t, defs in _raw.items() for sd in defs]
    status_choices.sort(key=lambda x: (x[1] or "").lower())
    
    type_status_colors = {t: {sd.key: sd.color_light for sd in defs if getattr(sd, "color_light", None)} for t, defs in _raw.items()}

    # Header Datum
    header_today = format_local_weekday_de(datetime.now(local_tz)) + ", " + datetime.now(local_tz).strftime("%d.%m.%Y")

    ctx = {
        "header_today": header_today,
        "request": request,
        "rows": rows,
        "type_status_options": type_status_options,
        "type_status_colors": type_status_colors,
        "status_choices": status_choices,
        "current_range": rng,
        "timedelta": timedelta,
        # Filter-Werte zurückgeben
        "q": q or "",
        "types": types or "",
        "status": status or "",
        "prio": prio or "",
        "show_private": show_private,
        "include_past": include_past,
        "tags": tags or "",
        "sort": sort or "",
        "dir": dir or "asc",
    }

    if is_htmx(request):
        return templates.TemplateResponse("_items_table.html", ctx)

    resp = templates.TemplateResponse("index.html", ctx)
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


# ====== Edit ======
@app.get("/items/{item_id}/edit", response_class=HTMLResponse)
def edit_item_page(
    item_id: str,
    request: Request,
    repo: DbRepository = Depends(get_repo),
    status=Depends(get_status),
):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")

    # Status-Optionen
    status_options = _status_options_for(status, it.type)
    if not status_options:
        status_options = [(it.status, _status_display(status, it.status))]

    # Recurrence-Form vorbereiten
    rrule_line = ""
    dtstart_local = ""
    exdates_local = ""
    if getattr(it, "recurrence", None) and it.recurrence.rrule_string:
        for line in it.recurrence.rrule_string.splitlines():
            if line.startswith("DTSTART:"):
                try:
                    dt = datetime.strptime(line.split(":", 1)[1], "%Y%m%dT%H%M%SZ").replace(tzinfo=ZoneInfo("UTC"))
                    dtstart_local = format_local(dt)
                except Exception:
                    pass
            if line.startswith("RRULE:"):
                rrule_line = line.split(":", 1)[1]
    if getattr(it, "recurrence", None) and it.recurrence.exdates_utc:
        exdates_local = ", ".join(format_local(d) for d in it.recurrence.exdates_utc)

    # Aktuelle Status-Farbe
    status_color = None
    if hasattr(status, "get_definition"):
        sd = status.get_definition(it.status)
        if sd and getattr(sd, "color_light", None):
            status_color = sd.color_light

    # Querystring für „Zurück“-Link sauber bauen
    back_qs = urlencode(list(request.query_params.multi_items()))

    resp = templates.TemplateResponse("edit.html", {
        "request": request,
        "it": it,
        "status_options": status_options,
        "dtstart_local": dtstart_local,
        "rrule_line": rrule_line,
        "exdates_local": exdates_local,
        "status_color": status_color,
        "back_qs": back_qs,
    })
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


URL_RE = re.compile(r"(https?://\S+)", re.IGNORECASE)

def _extract_links_from_text(text: str) -> list[str]:
    raw = URL_RE.findall(text or "")
    if not raw:
        return []

    TRAILING_ESCAPES = ("\\n", "\\r", "\\t")  # literal Backslash + Buchstabe
    TRAILING_PUNCT   = (".", ",", ";", ":", "!", "?", ")", "]", "}", "\"", "’", "”", "'")

    def trim_trailing(u: str) -> str:
        # Erst Whitespace an den Enden entfernen (falls vorhanden)
        u = (u or "").strip()
        # Iterativ: erst Escape-Suffixe, dann Satzzeichen kappen
        changed = True
        while changed and u:
            changed = False
            # literal-Escapes wie "\n", "\r", "\t" am Ende entfernen
            for esc in TRAILING_ESCAPES:
                if u.endswith(esc):
                    u = u[: -len(esc)]
                    changed = True
            # danach typische Satzzeichen am Ende entfernen
            while u and u[-1] in TRAILING_PUNCT:
                u = u[:-1]
                changed = True
        return u

    cleaned = [trim_trailing(u) for u in raw]
    seen = set()
    out = []
    for u in cleaned:
        if not u or u in seen:
            continue
        seen.add(u)
        out.append(u)
    return out

# typgerechte Statusprüfung
def valid_status_for_type(item_type: str, status_key: str) -> bool:
    if status_key is None:
        return True  # nichts ändern
    if item_type == "task":
        return status_key.startswith("TASK_")
    if item_type == "appointment":
        return status_key.startswith("APPOINTMENT_")
    if item_type == "event":
        return status_key.startswith("EVENT_")
    if item_type == "reminder":
        return status_key.startswith("REMINDER_") or (status_key == "")
    return False


@app.post("/items/{item_id}/edit")
async def edit_item_submit(
    item_id: str,
    request: Request,
    repo: DbRepository = Depends(get_repo),
    status=Depends(get_status),

    # Form-Bindung: alle optional, damit Partial-Updates funktionieren
    name: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    status_key: Optional[str] = Form(None),
    due: Optional[str] = Form(None),
    start_local: Optional[str] = Form(None),
    end_local: Optional[str] = Form(None),
    dtstart_local: Optional[str] = Form(None),
    rrule_line: Optional[str] = Form(None),
    exdates_local: Optional[str] = Form(None),
    is_private: Optional[str] = Form(None),
    is_all_day: Optional[str] = Form(None),
    tags: Optional[str] = Form(None),
    priority: Optional[str] = Form(None),
):
    it = repo.get(item_id)

    if not it:
        raise HTTPException(404, "Item nicht gefunden")

    # Effektive Werte bilden (Merge mit bestehendem Item), erst dann validieren
    eff_name = (name if name is not None else getattr(it, "name", "")) or ""
    eff_name = eff_name.strip()

    eff_description = (description if description is not None else getattr(it, "description", None))
    eff_description = (eff_description or "").strip() or None

    # Status normalisieren: wenn nichts geliefert, aktuellen übernehmen
    requested_status_key = status_key if (status_key or "") != "" else getattr(it, "status", None)
    requested_status_key = status_svc.normalize_input(requested_status_key or "", it.type) if requested_status_key else requested_status_key

    # Zeiten: nur parsen, wenn Feld geliefert; sonst bestehende Werte verwenden
    if it.type == "task":
        eff_due = _parse_local_dt(due) if (due or "").strip() else getattr(it, "due_utc", None)
        eff_start = getattr(it, "start_utc", None)
        eff_end = getattr(it, "end_utc", None)
    elif it.type in ("appointment", "event"):
        eff_start = _parse_local_dt(start_local) if (start_local or "").strip() else getattr(it, "start_utc", None)
        eff_end = _parse_local_dt(end_local) if (end_local or "").strip() else getattr(it, "end_utc", None)
        if eff_start and eff_end and eff_end < eff_start:
            eff_end = eff_start + timedelta(hours=1)
        eff_due = getattr(it, "due_utc", None)
    elif it.type == "reminder":
        eff_due = _parse_local_dt(due) if (due or "").strip() else getattr(it, "reminder_utc", None)
        eff_start = getattr(it, "start_utc", None)
        eff_end = getattr(it, "end_utc", None)
    else:
        eff_due = getattr(it, "due_utc", None)
        eff_start = getattr(it, "start_utc", None)
        eff_end = getattr(it, "end_utc", None)

    # Recurrence: nur neu setzen, wenn Eingaben vorhanden; sonst beibehalten
    new_rrule_string, new_exdates_utc = _normalize_rrule_input(dtstart_local or "", rrule_line or "", exdates_local or "")
    if (dtstart_local or rrule_line or exdates_local):
        eff_recurrence = _build_recurrence(new_rrule_string, new_exdates_utc)
    else:
        eff_recurrence = getattr(it, "recurrence", None)

    # Vorvalidierung auf Basis der effektiven Eingaben
    messages = _validate_edit_input(
        it, status, eff_name, requested_status_key,
        due, start_local, end_local,
        dtstart_local, rrule_line, exdates_local,
    )
    if messages:
        html = templates.get_template("_alerts.html").render({"messages": messages})
        return HTMLResponse(content=html, status_code=422)

    # Status-Transition prüfen
    old_key = getattr(it, "status", None)
    new_key = status_svc.normalize_input(requested_status_key or old_key, getattr(it, "type", None)) if (requested_status_key or old_key) else None
    is_recurring = bool(eff_recurrence)

    # ok, reason = status_svc.validate_transition(old_key, new_key, is_recurring=is_recurring)
    # if not ok:
    #     html = templates.get_template("_alerts.html").render({"messages": [reason or "Statuswechsel nicht erlaubt."]})
    #    return HTMLResponse(content=html, status_code=422)

    # Tags deduplizieren
    if tags is not None:
        tags_list = [t.strip() for t in tags.split(",") if t.strip()]
        tags_tuple = tuple(dict.fromkeys(tags_list))
    else:
        # Bestehende Tags beibehalten
        tags_tuple = getattr(it, "tags", None)

    # Priority nur ändern, wenn explizit gesendet
    if priority is not None:
        try:
            pv = int(priority)
            if 0 <= pv <= 5:
                prio_val = pv
            else:
                # Außerhalb Bereich -> bestehenden Wert behalten
                prio_val = getattr(it, "priority", None)
        except (ValueError, TypeError):
            # Nicht parsebar -> bestehenden Wert behalten
            prio_val = getattr(it, "priority", None)
    else:
        # Nicht gesendet -> bestehenden Wert behalten
        prio_val = getattr(it, "priority", None)

    # Privacy nur ändern, wenn explizit gesendet
    if is_private is not None:
        priv_bool = bool(int(is_private))
    else:
        # Nicht gesendet -> bestehenden Wert behalten
        priv_bool = bool(getattr(it, "is_private", False))

    # All-Day Flag
    if is_all_day is not None:
        allday_bool = bool(int(is_all_day))
    else:
        allday_bool = bool(getattr(it, "is_all_day", False))

    # Payload zusammenbauen
    payload = {
        **it.__dict__,
        "name": eff_name,
        "description": eff_description,
        "status": (requested_status_key or getattr(it, "status", None)),
        "is_private": priv_bool,
        "tags": tags_tuple,
        "priority": prio_val,
        "recurrence": eff_recurrence,
    }
    if it.type == "task":
        payload["due_utc"] = eff_due
    elif it.type in ("appointment", "event"):
        payload["start_utc"] = eff_start
        payload["end_utc"] = eff_end
        payload["is_all_day"] = allday_bool
    elif it.type == "reminder":
        payload["reminder_utc"] = eff_due

    # Links aus Beschreibung extrahieren (non-fatal)
    try:
        new_links = _extract_links_from_text(eff_description or "")
        if new_links:
            cur_links = list(getattr(it, "links", []) or [])
            cur_set = set(cur_links)
            added = False
            for u in new_links:
                if u not in cur_set:
                    cur_links.append(u)
                    cur_set.add(u)
                    added = True
            if added:
                payload["links"] = tuple(cur_links) if isinstance(getattr(it, "links", ()), tuple) else cur_links
    except Exception:
        pass

    # Auto-Finalisierung (nur non-recurring)
    if it.type in ("appointment", "event"):
        payload_item = {}
        if eff_end:  payload_item["end"] = eff_end
        if eff_start: payload_item["start"] = eff_start
        suggested = status_svc.auto_adjust_appointment_status(payload_item, now=now_utc())
        if suggested:
            payload["status"] = suggested

    # Persistieren
    it2 = it.__class__(**payload)

    repo.upsert(it2)
    repo.conn.commit()

    # HTMX-Teilantwort
    if is_htmx(request):
        # Prüfe ob von Listenansicht (Target ist eine Zeile)
        target = request.headers.get("HX-Target", "")
        
        if target.startswith("row-"):
            # Von Listenansicht: Ganze Zeile zurückgeben
            # Hole benötigte Kontextdaten
            # Im HTMX-Teil (Zeilen-Render):
            _raw = {t: status_svc.get_options_for(t) for t in ("task","reminder","appointment","event")}
            type_status_options = {t: [(sd.key, sd.display_name) for sd in defs] for t, defs in _raw.items()}

            row_html = templates.get_template("_items_table.html").render({
                "request": request,
                "rows": [(it2, [], None, None)],
                "type_status_options": type_status_options,
                "type_status_colors": {
                    t: {sd.key: sd.color_light for sd in defs if getattr(sd, "color_light", None)}
                    for t, defs in _raw.items()
                },
                "timedelta": timedelta,
            })
            
            # Extrahiere nur die TR aus dem gerenderten HTML
            import re
            match = re.search(r'(<tr[^>]*id="row-[^"]*"[^>]*>.*?</tr>)', row_html, re.DOTALL)
            if match:
                return HTMLResponse(content=match.group(1), status_code=200)
            else:
                # Fallback: leere Response
                return Response(status_code=204)
        
        # Von Detailansicht: Links und Occurrences aktualisieren
        n_raw = request.query_params.get("n")
        try:
            n = int(n_raw) if n_raw is not None else 10
        except Exception:
            n = 10
        n = max(1, min(n, 10))

        occs = _expand_next(it2, start_dt=now_utc(), max_count=n) or []
        links_html = templates.get_template("_links_block.html").render({
            "request": request,
            "it": it2,
        })
        occs_html = templates.get_template("_occurrences.html").render({
            "request": request,
            "occs": occs,
            "it": it2,
            "disp_start": None,
            "disp_end": None,
            "timedelta": timedelta,
        })
        html = f"""
    <div id="links-block" hx-swap-oob="true">
    {links_html}
    </div>
    <div id="occurrences" hx-swap-oob="true">
    {occs_html}
    </div>
    """
        return HTMLResponse(content=html, status_code=200)

def _parse_rrule_parts(rrule_str: str):
    parts = {}
    for token in (rrule_str or "").upper().split(";"):
        if "=" in token:
            k, v = token.split("=", 1)
            parts[k.strip()] = v.strip()
    return parts

def _occ_sort_key(occ):
    return (getattr(occ, "start_utc", None) or 
            getattr(occ, "due_utc", None) or 
            getattr(occ, "end_utc", None))

def _expand_next(it, start_dt, max_count: int = 10):
    """
    Liefert bis zu max_count Occurrences ab start_dt.
    Vergrößert das Auswertefenster iterativ, bis genügend Treffer vorhanden sind.
    """
    rec = getattr(it, "recurrence", None)
    rrule_str = (getattr(rec, "rrule_string", None) or "") if rec else ""
    parts = _parse_rrule_parts(rrule_str)

    freq = parts.get("FREQ", "DAILY")
    try:
        interval = int(parts.get("INTERVAL", "1"))
    except ValueError:
        interval = 1

    # Basisfenster in Tagen je Frequenz
    unit_days = 1
    if freq == "DAILY":
        unit_days = 1
    elif freq == "WEEKLY":
        unit_days = 7
    elif freq == "MONTHLY":
        unit_days = 30
    elif freq == "YEARLY":
        unit_days = 365

    # Intelligentere Startwert-Berechnung
    # Ziel: Mindestens max_count + Puffer im ersten Versuch
    safety_buffer = 1.5
    initial_horizon_days = max(
        365,  # Minimum: 1 Jahr
        int(max_count * interval * unit_days * safety_buffer)
    )

    # Höheres Iterations-Limit
    hard_cap_days = 365 * 20  # Bis zu 20 Jahre
    hard_cap_iters = 15  # Mehr Iterationen erlauben

    horizon_days = initial_horizon_days
    iters = 0
    results = []

    while iters < hard_cap_iters and horizon_days <= hard_cap_days:
        win_end = start_dt + timedelta(days=horizon_days)
        seg = expand_item(it, start_dt, win_end) or []
        seg_sorted = sorted(seg, key=_occ_sort_key)
        results = [s for s in seg_sorted if (_occ_sort_key(s) is not None)]
        
        if len(results) >= max_count:
            break
        
        # Aggressivere Expansion bei wenigen Treffern
        if len(results) < max_count // 2:
            # Sehr wenige Treffer → 3x Multiplikator
            horizon_days = int(horizon_days * 3)
        else:
            # Fast genug → 1.5x Multiplikator
            horizon_days = int(horizon_days * 1.5)
        
        iters += 1

    return results[:max_count]


@app.get("/items/{item_id}/occurrences", response_class=HTMLResponse)
def occurrences(request: Request, item_id: str, n: int = 10, repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    n = max(1, min(int(n or 10), 10))
    now = now_utc()

    # Serien?
    rec = getattr(it, "recurrence", None)
    has_recur = False
    if rec:
        rs = (getattr(rec, "rrule_string", None) or getattr(rec, "rrulestring", None) or "").strip()
        ex = getattr(rec, "exdates_utc", None)
        has_recur = bool(rs) or bool(ex and len(ex) > 0)

    occs = []
    if has_recur:
        occs = _expand_next(it, start_dt=now, max_count=n) or []
    else:
        # YEARLY-Semantik (Geburtstag) → immer nächstes Vorkommen
        if getattr(it, "type", "") == "event":
            if is_birthday(it):
                s, e = next_or_display_occurrence(it, now=now)
                if s:
                    occs = [{"item_type":"event","start_utc":s,"end_utc":e}]
            else:
                # Nicht-Serie: nur aktuelles
                s = getattr(it, "start_utc", None)
                e = getattr(it, "end_utc", None)
                if s or e:
                    occs = [{"item_type":"event","start_utc":s,"end_utc":e}]
        elif getattr(it, "type", "") == "appointment":
            s = getattr(it, "start_utc", None)
            e = getattr(it, "end_utc", None)
            if s or e:
                occs = [{"item_type":"appointment","start_utc":s,"end_utc":e}]
        elif getattr(it, "type", "") in ("task","reminder"):
            d = getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None)
            if d:
                occs = [{"item_type":getattr(it,"type"),"due_utc":d}]

    return templates.TemplateResponse("_occurrences.html", {"request": request, "occs": occs, "timedelta": timedelta, "it": it, "disp_start": None, "disp_end": None})

# ====== Statuswechsel (zentraler StatusService) ======
@app.post("/items/{item_id}/status")
def change_status(
    request: Request,
    item_id: str,
    new_status: str = Form(...),
    repo: DbRepository = Depends(get_repo),
):
    it = repo.get(item_id)
    if not it:
        return Response(status_code=204) if is_htmx(request) else RedirectResponse("/", status_code=303)

    # 1) Eingabewert typbewusst normalisieren (Altformen + Präfix-Garantie)
    requested = status_svc.normalize_input(new_status, it.type)

    # 2) Erlaubte Optionen vom Service holen und auf Keys abbilden
    options = status_svc.get_options_for(getattr(it, "type", None))
    allowed_keys = {opt.key for opt in options}

    # 3) Fallback: Label-Match erlauben (Case-insensitive), falls Nutzer ein Label gesendet hat
    if requested not in allowed_keys:
        ns_l = (new_status or "").strip().lower()
        for opt in options:
            if (opt.display_name or "").strip().lower() == ns_l:
                requested = opt.key
                break

    # 4) Harte Allowed-Prüfung
    if allowed_keys and requested not in allowed_keys:
        if is_htmx(request):
            html = templates.get_template("_alerts.html").render(
                {"messages": [f"Ungültiger Status für {it.type}."]}
            )
            return HTMLResponse(content=html, status_code=422)
        return RedirectResponse("/", status_code=303)

    # 5) Übergangsvalidierung (inkl. Serien-/Reminder-Sonderfälle im Manager)
    is_recurring = bool(getattr(it, "recurrence", None) or getattr(it, "rrule_string", None))
    ok, reason = status_svc.validate_transition(getattr(it, "status", None), requested, is_recurring=is_recurring)

    if not ok:
        if is_htmx(request):
            html = templates.get_template("_alerts.html").render(
                {"messages": [reason or "Statuswechsel nicht erlaubt."]}
            )
            return HTMLResponse(content=html, status_code=422)
        return RedirectResponse("/", status_code=303)

    # 6) Schreiben nur bei tatsächlicher Änderung
    if requested != getattr(it, "status", None):
        it2 = it.__class__(**{**it.__dict__, "status": requested})
        repo.upsert(it2)
        repo.conn.commit()

    return Response(status_code=204) if is_htmx(request) else RedirectResponse("/", status_code=303)


# ====== Teilansicht Tabelle ======
@app.get("/items/table", response_class=HTMLResponse)
def items_table(request: Request, repo: DbRepository = Depends(get_repo)):
    items = repo.list_all()
    rows = []
    now_dt = datetime.now(timezone.utc)
    for it in items:
        if getattr(it, "type", "") in ("appointment","event"):
            disp_start, disp_end = next_or_display_occurrence(it, now=now_dt)
        else:
            disp_start = getattr(it, "due_utc", None) or getattr(it, "start_utc", None)
            disp_end   = getattr(it, "end_utc", None) or getattr(it, "reminder_utc", None)
        occs = _expand_next(it, start_dt=now_dt, max_count=1) or []
        rows.append((it, occs, disp_start, disp_end))

    _raw = {t: status_svc.get_options_for(t) for t in ("task","reminder","appointment","event")}
    type_status_options = {t: [(sd.key, sd.display_name) for sd in defs] for t, defs in _raw.items()}
    type_status_colors = {t: {sd.key: sd.color_light for sd in defs if getattr(sd, "color_light", None)} for t, defs in _raw.items()}

    return templates.TemplateResponse("_items_table.html", {
        "request": request,
        "rows": rows,
        "type_status_options": type_status_options,
        "type_status_colors": type_status_colors,
    })


# ====== Rename / Due / Start-End / Snooze ======
@app.post("/items/{item_id}/rename")
def rename_item(item_id: str, name: str = Form(...), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")
    name = (name or "").strip()
    if not name:
        return HTMLResponse(templates.get_template("_alerts.html").render({"messages": ["Name darf nicht leer sein."]}), status_code=422)
    if len(name) > 200:
        return HTMLResponse(templates.get_template("_alerts.html").render({"messages": ["Name ist zu lang."]}), status_code=422)
    it2 = it.__class__(**{**it.__dict__, "name": name})
    repo.upsert(it2)
    repo.conn.commit()
    return Response(status_code=204)


@app.post("/items/{item_id}/due")
def set_due(request: Request, item_id: str, due: str = Form(...), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it or it.type != "task":
        return RedirectResponse("/", status_code=303)
    if not (due or "").strip():
        return RedirectResponse("/", status_code=303)

    dt_utc = _parse_local_dt(due)
    it2 = it.__class__(**{**it.__dict__, "due_utc": dt_utc})
    repo.upsert(it2)
    repo.conn.commit()

    if is_htmx(request):
        now_dt = now_utc()
        disp_start = _aware(dt_utc)
        disp_end = None
        occs = _expand_next(it2, start_dt=now_dt, max_count=3) or []

        return templates.TemplateResponse(
            "_occurrences.html",
            {"request": request, "it": it2, "occs": occs, "disp_start": disp_start, "disp_end": disp_end, "timedelta": timedelta,},
        )

    return RedirectResponse("/", status_code=303)


@app.post("/items/{item_id}/start_end")
def set_start_end(
    request: Request,
    item_id: str,
    start_local: str = Form(""),
    end_local: str = Form(""),
    repo: DbRepository = Depends(get_repo),
):
    it = repo.get(item_id)
    if not it or it.type not in ("appointment", "event"):
        return RedirectResponse("/", status_code=303)

    if not (start_local or "").strip() and not (end_local or "").strip():
        return RedirectResponse("/", status_code=303)

    s_utc = _parse_local_dt(start_local) if (start_local or "").strip() else getattr(it, "start_utc", None)
    e_utc = _parse_local_dt(end_local) if (end_local or "").strip() else getattr(it, "end_utc", None)

    if s_utc and e_utc and e_utc < s_utc:
        e_utc = s_utc + timedelta(hours=1)

    it2 = it.__class__(**{**it.__dict__, "start_utc": s_utc, "end_utc": e_utc})
    repo.upsert(it2)
    repo.conn.commit()

    if is_htmx(request):
        now_dt = now_utc()
        ds, de = next_or_display_occurrence(it2, now=now_dt)
        disp_start = _aware(ds)
        disp_end   = _aware(de)
        occs = _expand_next(it2, start_dt=now_dt, max_count=3) or []

        return templates.TemplateResponse(
            "_occurrences.html",
            {"request": request, "it": it2, "occs": occs, "disp_start": disp_start, "disp_end": disp_end, "timedelta": timedelta,},
        )

    return RedirectResponse("/", status_code=303)


@app.post("/items/{item_id}/snooze")
def snooze(request: Request, item_id: str, minutes: int = Form(10), until_local: str = Form(""), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if it and it.type == "reminder":
        new_ts = None
        if (until_local or "").strip():
            new_ts = _parse_local_dt(until_local)
        if not new_ts:
            new_ts = (it.reminder_utc or now_utc()) + timedelta(minutes=int(minutes))
        it2 = it.__class__(**{**it.__dict__, "reminder_utc": new_ts})
        repo.upsert(it2)
        repo.conn.commit()
        if is_htmx(request):
            now_dt = now_utc()
            disp_start = _aware(new_ts)
            disp_end = None
            occs = _expand_next(it2, start_dt=now_dt, max_count=3) or []

            return templates.TemplateResponse(
                "_occurrences.html",
                {"request": request, "it": it2, "occs": occs, "disp_start": disp_start, "disp_end": disp_end, "timedelta": timedelta,},
            )
    if is_htmx(request):
        return Response(status_code=204)
    return RedirectResponse("/", status_code=303)


# ====== Delete / New ======
@app.post("/items/{item_id}/delete")
def delete_item(request: Request, item_id: str, repo: DbRepository = Depends(get_repo)):
    ok = repo.delete(item_id)
    if ok:
        repo.conn.commit()
    if is_htmx(request):
        return hx_refresh()
    return RedirectResponse("/", status_code=303)


@app.post("/items/delete_selected")
def delete_selected(ids: Annotated[List[str], Form(...)], repo: DbRepository = Depends(get_repo)):
    try:
        repo.conn.execute("BEGIN")
        for iid in ids:
            repo.delete(iid)
        repo.conn.commit()
    except Exception:
        repo.conn.rollback()
        raise
    return RedirectResponse("/", status_code=303)


@app.post("/items/new")
def create_item(
    request: Request,
    name: str = Form(...),
    item_type: str = Form(...),
    repo: DbRepository = Depends(get_repo),
):
    nid = str(uuid.uuid4())
    item_type = (item_type or "").strip().lower()
    name = (name or "").strip()

    def htmx_error(msg: str):
        if request.headers.get("HX-Request") == "true":
            # Antwort eignet sich für Einblendung in #alerts oder einen globalen Listener
            return HTMLResponse(f'<div class="alert alert-error">{msg}</div>', status_code=422)
        raise HTTPException(status_code=422, detail=msg)
    
    # 1) Frühe Validierung
    valid_types = {"task", "reminder", "appointment", "event"}

    item_type = (item_type or "").strip().lower()
    name = (name or "").strip()

    errors = []
    if not item_type or item_type not in valid_types:
        errors.append("Bitte einen gültigen Typ wählen. ")
    if not name:
        errors.append("Bitte einen Namen eingeben. ")

    if errors:
        # HTMX-Client: kompakter HTML-Alert, ansonsten HTTP 422
        if request.headers.get("HX-Request") == "true":
            html = '<div class="alert alert-error">' + '<br>'.join(errors) + '</div>'
            # Optional: gezielt in #alerts einfügen per hx-target auf Clientseite
            return HTMLResponse(content=html, status_code=422)
        raise HTTPException(status_code=422, detail="; ".join(errors))

    # 2) Default-Status dynamisch aus Status-Service
    opts = status_svc.get_options_for(item_type=item_type)
    non_terminal = [sd for sd in opts if not sd.is_terminal]
    if non_terminal:
        default_status = non_terminal[0].key
    elif opts:
        default_status = opts[0].key
    else:
        # Kein definierter Status für diesen Typ -> 422
        if request.headers.get("HX-Request") == "true":
            return HTMLResponse('<div class="alert alert-error">Für den gewählten Typ ist kein Status definiert.</div>', status_code=422)
        raise HTTPException(status_code=422, detail="Für den gewählten Typ ist kein Status definiert.")

    # 3) Fabriken – klare Typzuordnung, kein Fallback auf „task“
    def mk_task():
        return Task(id=nid, type="task", name=name, status=default_status,
                    is_private=False, due_utc=None, recurrence=None)

    def mk_reminder():
        return Reminder(id=nid, type="reminder", name=name, status=default_status,
                        is_private=False, reminder_utc=None, recurrence=None)

    def mk_appointment(_type="appointment"):
        return Appointment(id=nid, type=_type, name=name, status=default_status,
                        is_private=False, start_utc=None, end_utc=None,
                        is_all_day=False, recurrence=None)

    def mk_event():
        if Event:
            return Event(id=nid, type="event", name=name, status=default_status,
                        is_private=False, start_utc=None, end_utc=None,
                        is_all_day=False, recurrence=None)
        return mk_appointment(_type="event")

    factories = {
        "task": mk_task,
        "reminder": mk_reminder,
        "appointment": mk_appointment,
        "event": mk_event,
    }

    it = factories[item_type]()

    # 4) Optional: Status-Schlüssel absichern, falls Definitionen angepasst wurden
    try:
        norm_key = status_svc.normalize_input(default_status, item_type=item_type)
        if norm_key and norm_key != it.status:
            it = it.__class__(**{**it.__dict__, "status": norm_key})
    except Exception:
        pass

    # 5) Persistenz
    repo.upsert(it)
    repo.conn.commit()

    edit_url = f"/items/{nid}/edit"

    # 6) Antwort
    if is_htmx(request):
        resp = Response(status_code=204)
        resp.headers["HX-Redirect"] = edit_url
        return resp
    return RedirectResponse(edit_url, status_code=303)


@app.post("/items/{item_id}/copy")
def copy_item(item_id: str, request: Request, repo: DbRepository = Depends(get_repo)):
    try:
        new_obj = repo.copy_item(item_id)
        repo.conn.commit()
    except ValueError:
        if request.headers.get("HX-Request") == "true":
            return HTMLResponse('<div class="alert alert-error">Item nicht gefunden.</div>', status_code=404)
        raise HTTPException(status_code=404, detail="Item nicht gefunden.")

    edit_url = f"/items/{new_obj.id}/edit?copied=1"
    if request.headers.get("HX-Request") == "true":
        resp = Response(status_code=204)
        resp.headers["HX-Redirect"] = edit_url
        return resp
    return RedirectResponse(edit_url, status_code=303)


# ====== ICS-Export (Item & Global & Auswahl) ======
@app.get("/items/{item_id}/export.ics")
def export_item_ics(item_id: str, repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(status_code=404, detail="Item nicht gefunden")

    # Einzel-Component erzeugen (BEGIN:VEVENT ... END:VEVENT)
    comp = to_ics(it, alarm_min=10)  # enthält bereits UID/PRIORITY [attached_file:15][attached_file:3]

    # VCALENDAR um die Komponente bauen
    cal_lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//TaskManager//DE",
        "CALSCALE:GREGORIAN",
        comp,
        "END:VCALENDAR",
        "",
    ]
    body = "\n".join(cal_lines)

    # Dateiname robust ableiten
    base = (it.name or "item").strip().replace(" ", "_")
    # Optional: Nicht-ASCII/Problemzeichen entfernen
    safe = "".join(ch for ch in base if ch.isalnum() or ch in ("_", "-", "."))
    filename = f"{safe or 'item'}.ics"

    return StreamingResponse(
        iter([body.encode("utf-8")]),
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )

@app.get("/export.ics")
def export_ics(
    request: Request,
    ids: list[str] = Query(default=None),  # Mehrfach-IDs ?ids=...&ids=...
    repo: DbRepository = Depends(get_repo)
):
    # 1) Auswahl aus Liste (ids vorhanden)
    if ids:
        items = [x for x in (repo.get(i) for i in ids) if x]
    else:
        # 2) Edit-Kontext über Referer erkennen
        ref = (request.headers.get("Referer") or "").lower()
        m = re.search(r"/items/([0-9a-f-]+)/edit", ref)
        if m:
            it = repo.get(m.group(1))
            items = [it] if it else []
        else:
            # 3) Fallback: alle Items
            items = repo.list_all()

    comps = [to_ics(it, alarm_min=10) for it in items]  # STATUS/DT*/DUE/PRIORITY/CREATED/LAST-MODIFIED
    head = (
        "BEGIN:VCALENDAR\n"
        "VERSION:2.0\n"
        "PRODID:-//TaskManager//DE\n"
        "CALSCALE:GREGORIAN\n"
        "METHOD:PUBLISH\n"
    )
    body = "\n".join(comps) + "\n"
    tail = "END:VCALENDAR\n"
    ics_text = head + body + tail

    fname = "items.ics" if len(items) != 1 else f"{(items[0].name or 'item').strip().replace(' ','_')}.ics"
    return StreamingResponse(
        iter([ics_text.encode("utf-8")]),
        media_type="text/calendar; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

@app.post("/export_selected")
def export_selected(
    ids: Annotated[Optional[List[str]], Form()] = None,
    q: str = Form(""),
    types: str = Form(""),
    status_keys: str = Form(""),
    show_private: int = Form(0),
    tags: str = Form(""),
    include_past: int = Form(0),
    repo: DbRepository = Depends(get_repo),
    status=Depends(get_status),
):
    # Alle Items laden und mit der bestehenden Filterlogik eingrenzen
    items = repo.list_all()

    # Optional: vorhandenen Filter-Service verwenden, falls vorhanden
    try:
        items = filter_items(
            items=items,
            q=q,
            types=types,
            status_keys=status_keys,
            show_private=bool(int(show_private or 0)),
            tags=tags,
            include_past=bool(int(include_past or 0)),
            status_service=status,
        )
    except Exception:
        # Fallback: nur minimal nach q filtern
        q_norm = (q or "").strip().lower()
        if q_norm:
            items = [it for it in items if q_norm in (getattr(it, "name", "") or "").lower()]

    # Selektion auf ausgewählte IDs einschränken, wenn übergeben
    if ids:
        idset = set(ids)
        items = [it for it in items if getattr(it, "id", None) in idset]

    # ICS-Kalender zusammenbauen
    comps = []
    for it in items:
        try:
            comps.append(to_ics(it, alarm_min=10))  # sollte BEGIN:VEVENT ... END:VEVENT enthalten
        except Exception:
            # Einzelne Items, die nicht gemappt werden können, überspringen
            continue

    cal_header = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//TaskManager//DE",
        "CALSCALE:GREGORIAN",
    ]
    cal_body = "\n".join(cal_header) + "\n" + "\n".join(comps) + "\nEND:VCALENDAR\n"

    # Dateiname abhängig von Auswahl vs. Filter
    fname = "export_selection.ics" if ids else "export_filtered.ics"

    # StreamingResponse mit FastAPI ausliefern
    return StreamingResponse(
        iter([cal_body.encode("utf-8")]),
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "Cache-Control": "no-store",
        },
    )


# ===== Dashboard-Hilfsfunktionen =====

def is_birthday(item) -> bool:
    """True, wenn Event ausschließlich mit Tag 'geburtstag' versehen ist."""
    if getattr(item, "type", None) != "event":
        return False
    tags = getattr(item, "tags", None) or ()
    tags_norm = [str(t).strip().lower() for t in tags if t is not None]
    return len(tags_norm) == 1 and tags_norm[0] == "geburtstag"


def format_dashboard_time(dt: datetime, context: str, local_tz=None) -> str:
    """
    Formatiert Datum/Zeit kontextabhängig für Dashboard.
    
    Args:
        dt: datetime (UTC oder aware)
        context: "series", "next_events", "calendar", "today", "next_48h", "next_7d", "no_date"
        local_tz: ZoneInfo für Konvertierung (Standard: Europe/Berlin)
    
    Returns:
        Formatierter String
    """
    if not dt:
        return ""
    
    if local_tz is None:
        local_tz = ZoneInfo("Europe/Berlin")
    
    # Sicherstellen dass dt timezone-aware ist
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    
    local_dt = dt.astimezone(local_tz)
    
    if context == "series":
        # TT.MM.
        return local_dt.strftime("%d.%m.")
    
    elif context == "next_events":
        # TT.MM.JJJJ HH:mm
        return local_dt.strftime("%d.%m.%Y %H:%M")
    
    elif context in ("calendar", "today"):
        # HH:mm
        return local_dt.strftime("%H:%M")
    
    elif context == "next_48h":
        # TT.MM HH:mm
        return local_dt.strftime("%d.%m %H:%M")
    
    elif context == "next_7d":
        # Wochentag TT.MM. HH:mm
        weekdays = ["Mo.", "Di.", "Mi.", "Do.", "Fr.", "Sa.", "So."]
        weekday = weekdays[local_dt.weekday()]
        return f"{weekday} {local_dt.strftime('%d.%m. %H:%M')}"
    
    elif context == "no_date":
        # TT.MM.JJJJ (Änderungsdatum)
        return local_dt.strftime("%d.%m.%Y")
    
    return local_dt.strftime("%d.%m.%Y %H:%M")


def get_priority_class(item) -> str:
    """Gibt CSS-Klasse für Priorität zurück."""
    priority = getattr(item, "priority", 0)
    
    if priority >= 3:
        return "priority-high"
    elif priority == 2:
        return "priority-medium"
    elif priority == 1:
        return "priority-low"
    
    return ""


def is_overdue_item(item, now_dt: datetime) -> bool:
    """Prüft ob Item überfällig ist."""
    # Relevante Zeit je Typ
    if item.type in ("appointment", "event"):
        relevant_dt = getattr(item, "start_utc", None)
        end_dt = getattr(item, "end_utc", None)
        # Als überfällig nur werten, wenn komplett vorbei:
        if end_dt is not None:
            return (end_dt < now_dt) and not status_svc.is_terminal(item.status)
    elif item.type in ("task", "reminder"):
        relevant_dt = getattr(item, "due_utc", None) or getattr(item, "reminder_utc", None)
    else:
        relevant_dt = None

    if not relevant_dt:
        return False

    # Überfällig = Datum vergangen UND Status nicht terminal (zentral geprüft)
    is_past = relevant_dt < now_dt
    is_open = not status_svc.is_terminal(getattr(item, "status", None))

    return is_past and is_open


# ====== Import & Tags & Links ======
@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    return templates.TemplateResponse("import.html", {"request": request})

@app.post("/import", response_class=HTMLResponse)
async def import_upload(
    request: Request, 
        back_qs: str = Form(""), 
        file: UploadFile = File(...), 
        repo: DbRepository = Depends(get_repo),
):
    text = (await file.read()).decode("utf-8", errors="ignore")

    # 1) ICS parsen (liefert bereits angereicherte Items: description, tags, links, metadata, audit)
    from services.ics_import import import_ics
    items = import_ics(text)

    # 2) Optionale Deduplizierung anhand ICS-UID (wenn im metadata gesetzt)
    def find_by_ics_uid(uid: str):
        return repo.get_by_ics_uid(uid) if uid else None

    def default_status_for(it_type: str) -> str:
        if it_type == "task": return "TASK_OPEN"
        if it_type == "reminder": return "REMINDER_ACTIVE"
        if it_type == "appointment": return "APPOINTMENT_PLANNED"
        if it_type == "event": return "EVENT_SCHEDULED"
        return "TASK_OPEN"

    # 3) Normalisieren und transaktional schreiben
    try:
        repo.conn.execute("BEGIN")
        for it in items:
            payload = dict(it.__dict__)

            # ICS-UID direkt ins Payload übernehmen (falls Importer sie noch nicht gesetzt hätte)
            ics_uid = payload.get("ics_uid") or (payload.get("metadata") or {}).get("ics_uid")
            if ics_uid:
                payload["ics_uid"] = ics_uid

            # ID setzen falls leer, anhand ics_uid deduplizieren
            if not payload.get("id"):
                if payload.get("ics_uid"):
                    existing = find_by_ics_uid(payload["ics_uid"])
                    if existing:
                        payload["id"] = existing.id
                    else:
                        payload["id"] = str(uuid.uuid4())
                else:
                    payload["id"] = str(uuid.uuid4())

            # Status fallback, falls Mapping leer war
            if not payload.get("status"):
                payload["status"] = default_status_for(payload.get("type") or "")

            # Description trimmen
            if "description" in payload and payload["description"]:
                payload["description"] = (payload["description"] or "").strip() or None

            # Tags/Links zu Tupeln normalisieren
            if "tags" in payload and payload["tags"] is not None:
                payload["tags"] = tuple(dict.fromkeys([t.strip() for t in (payload["tags"] or []) if t and str(t).strip()]))
            if "links" in payload and payload["links"] is not None:
                payload["links"] = tuple(dict.fromkeys([l.strip() for l in (payload["links"] or []) if l and str(l).strip()]))

            # Metadata initialisieren und ICS UID persistieren, falls vorhanden
            md = dict(payload.get("metadata") or {})
            # Versuch, eine UID aus dem ICS zu übernehmen (dein import_ics sollte sie setzen; hier als Fallback)
            if "ics_uid" not in md and "uid" in md:
                md["ics_uid"] = md["uid"]
            payload["metadata"] = md

            # Audit:
            # - created_utc/last_modified_utc aus ICS bleiben, wenn gesetzt
            # - andernfalls setzt upsert last_modified_utc automatisch und created_utc beim ersten Insert
            # Keine weitere Logik hier nötig

            # Rekonstruktion in passendes Klassenobjekt
            cls = it.__class__
            enriched = cls(**payload)

            repo.upsert(enriched)

        repo.conn.commit()
    except Exception:
        repo.conn.rollback()
        raise

    # UX
    if is_htmx(request):
        return Response(status_code=204, headers={"HX-Redirect": f"/?{back_qs}" if back_qs else "/"})
    return RedirectResponse(f"/?{back_qs}" if back_qs else "/", status_code=303)


@app.get("/tags/suggest")
def tags_suggest(q: str = "", repo: DbRepository = Depends(get_repo)):
    items = repo.list_all()
    c = Counter(t for it in items for t in (getattr(it, "tags", []) or []) if t)
    ql = (q or "").strip().lower()
    out = [t for t,_ in c.most_common() if not ql or t.lower().startswith(ql)]
    return out[:20]


@app.post("/items/{item_id}/tags/add")
def tags_add(item_id: str, tag: str = Form(...), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")
    
    tag = (tag or "").strip()
    if not tag:
        return HTMLResponse(
            templates.get_template("_alerts.html").render({"messages": ["Tag darf nicht leer sein."]}), 
            status_code=422
        )
    
    # Prüfe ob Event + Tag "geburtstag" + start_utc → Zeige Bestätigungsdialog
    if (
        it.type == "event"
        and tag.lower() == "geburtstag"
        and getattr(it, "start_utc", None)
        and not (getattr(it.recurrence, "rrule_string", None) if it.recurrence else False)
    ):
        # Zeige Bestätigungsdialog
        from html import escape
        tag_escaped = escape(tag)
        return HTMLResponse(
            f"""
            <div id="birthday-confirm-dialog" style="padding: 16px; background: var(--color-bg-2); border: 1px solid var(--color-border); border-radius: var(--radius-base); margin-bottom: 16px;">
                <p style="margin: 0 0 12px 0; font-weight: var(--font-weight-medium);">Jährlicher Geburtstag?</p>
                <p style="margin: 0 0 16px 0; font-size: var(--font-size-sm); color: var(--color-text-secondary);">
                    Soll dieser Termin jährlich wiederholt werden?
                </p>
                <div style="display: flex; gap: 8px;">
                    <form method="post" action="/items/{item_id}/tags/add/confirm" style="display: inline;">
                        <input type="hidden" name="tag" value="{tag_escaped}">
                        <input type="hidden" name="create_rrule" value="yes">
                        <button type="submit" class="btn btn--primary btn--sm">Ja, jährlich</button>
                    </form>
                    <form method="post" action="/items/{item_id}/tags/add/confirm" style="display: inline;">
                        <input type="hidden" name="tag" value="{tag_escaped}">
                        <input type="hidden" name="create_rrule" value="no">
                        <button type="submit" class="btn btn--secondary btn--sm">Nein, einmalig</button>
                    </form>
                </div>
            </div>
            """,
            status_code=200,
        )
    
    # Normaler Fall: Tag direkt hinzufügen
    tags = list(dict.fromkeys([*(it.tags or []), tag]))
    it2 = it.__class__(**{**it.__dict__, "tags": tags})
    repo.upsert(it2)
    repo.conn.commit()
    return Response(status_code=204)


@app.post("/items/{item_id}/tags/add/confirm")
def tags_add_confirm(
    item_id: str,
    tag: str = Form(...),
    create_rrule: str = Form(...),
    repo: DbRepository = Depends(get_repo),
):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")
    
    tag = (tag or "").strip()
    tags = list(dict.fromkeys([*(it.tags or []), tag]))
    
    payload = {"tags": tags}
    
    # Wenn Benutzer "Ja" gewählt hat → RRULE erstellen
    if create_rrule == "yes" and it.type == "event":
        sutc = getattr(it, "start_utc", None)
        if sutc:
            from domain.models import Recurrence
            dtstart_str = sutc.strftime("%Y%m%dT%H%M%SZ")
            rrule_str = f"DTSTART:{dtstart_str}\nRRULE:FREQ=YEARLY;INTERVAL=1"
            payload["recurrence"] = Recurrence(rrule_string=rrule_str, exdates_utc=())
    
    it2 = it.__class__(**{**it.__dict__, **payload})
    repo.upsert(it2)
    repo.conn.commit()
    
    return Response(status_code=204)


@app.post("/items/{item_id}/tags/remove")
def tags_remove(item_id: str, tag: str = Form(...), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")
    tag = (tag or "").strip()
    tags = [t for t in (getattr(it, "tags", []) or []) if t != tag]
    it2 = it.__class__(**{**it.__dict__, "tags": tuple(tags)})
    repo.upsert(it2)
    repo.conn.commit()
    return Response(status_code=204)


@app.post("/items/{item_id}/links/add")
def links_add(item_id: str, url: str = Form(...), repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")
    url = (url or "").strip()
    if not url or not (url.startswith("http://") or url.startswith("https://")):
        html = templates.get_template("_alerts.html").render({"messages": ["Ungültige URL."]})
        return HTMLResponse(html, status_code=422)

    links = list(getattr(it, "links", ()) or [])
    if url not in links:
        links.append(url)
    it2 = it.__class__(**{**it.__dict__, "links": tuple(dict.fromkeys(links))})
    repo.upsert(it2)
    repo.conn.commit()

    html = templates.get_template("_links_block.html").render({"it": it2})
    return HTMLResponse(html, headers={"HX-Trigger": "links-updated"})

@app.post("/items/{item_id}/links/remove")
def links_remove(
    request: Request,
    item_id: str,
    url: str = Form(...),
    repo: DbRepository = Depends(get_repo),
):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")

    url = (url or "").strip()
    cur_links = list(getattr(it, "links", ()) or [])
    new_links = [u for u in cur_links if u != url]

    # Nur persistieren, wenn sich etwas geändert hat
    if len(new_links) != len(cur_links):
        it = it.__class__(**{**it.__dict__, "links": tuple(new_links) if isinstance(it.links, tuple) else new_links})
        repo.upsert(it)
        repo.conn.commit()

    # Erkennen, ob Aufruf aus der Edit-Ansicht kommt:
    # - Edit nutzt _links_block.html mit hx-target="#links_block" + hx-swap="outerHTML"
    # - Liste (Items-Tabelle) nutzt hx-swap="none" und entfernt clientseitig .link-row bei 204
    referer = (request.headers.get("Referer") or "").lower()
    is_edit_context = "/items/" in referer and "/edit" in referer

    if is_edit_context:
        # Edit: kompletten Links-Block neu rendern
        html = templates.get_template("_links_block.html").render({
            "request": request,
            "it": it,
        })
        # Optional: Erfolgsmeldung via Trigger (dein base.html hört auf links-updated)
        return HTMLResponse(html, headers={"HX-Trigger": "links-updated"}, status_code=200)

    # Liste: 204 -> dein JS entfernt .link-row selbst (siehe htmx:afterRequest in _items_table.html)
    return Response(status_code=204)

# ====== Dashboard ======
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    q: str | None = None,
    types: str | None = None,
    status_keys: str | None = None,
    status: str | None = None,
    show_private: int = 0,
    include_past: int = 0,
    tags: str | None = None,
    cal_weeks: int = 2,
    cal_week_offset: int = 0,
    repo: DbRepository = Depends(get_repo),
    sm=Depends(get_status),
):
    print("=== DASHBOARD DEBUG ===")
    print("raw query:", dict(request.query_params))
    print("q:", q, "types:", types, "status_keys:", status_keys, "status:", status)
    print("show_private:", show_private, "include_past:", include_past, "tags:", tags, "cal_week_offset:", cal_week_offset)
    print("=======================")

    berlin = ZoneInfo("Europe/Berlin")
    today = now_utc().astimezone(berlin)

    # Daten laden und Privatsphäre filtern
    items = repo.list_all()
    include_private = bool(int(show_private or 0))
    def _visible_by_privacy(it) -> bool:
        is_private = bool(getattr(it, "private", False))
        return include_private or not is_private
    items = [it for it in items if _visible_by_privacy(it)]
    print(f"[DB] total items: {len(items)}")

    # Auto-Finalisierung vergangener Termine/Ereignisse
    changed = False
    updated = []
    for it in items:
        if getattr(it, "type", "") in ("appointment", "event"):
            payload = {}
            if getattr(it, "end_utc", None): payload["end"] = it.end_utc
            elif getattr(it, "until", None): payload["until"] = it.until
            if getattr(it, "start_utc", None): payload["start"] = it.start_utc
            if payload:
                suggested = status_svc.auto_adjust_appointment_status(payload, now=now_utc())
                if suggested and suggested != getattr(it, "status", ""):
                    it2 = it.__class__(**{**it.__dict__, "status": suggested})
                    repo.upsert(it2)
                    updated.append(it2)
                    changed = True
    if changed:
        repo.conn.commit()
        id2obj = {it.id: it for it in items}
        for it2 in updated: id2obj[it2.id] = it2
        items = list(id2obj.values())

    # Quick-Übersichten
    by_type = Counter(getattr(it, "type", "") for it in items)
    by_status = Counter(getattr(it, "status", "") for it in items)

    # Zeitfenster (lokal)
    now_local = today
    start_today = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    end_today = now_local.replace(hour=23, minute=59, second=59, microsecond=999999)
    end_7d = (start_today + timedelta(days=7)).replace(hour=23, minute=59, second=59, microsecond=999999)

    def as_local(dt_utc: datetime | None) -> datetime | None:
        return dt_utc.astimezone(berlin) if dt_utc else None

    upcoming, overdue, upcoming_today, upcoming_next7, without_date, recurring_items = [], [], [], [], [], []

    win_end_48h = now_utc() + timedelta(hours=48)

    # Serien-Panel: „aktiv in dieser Woche“
    week_monday_local = today - timedelta(days=today.weekday())
    week_monday_local = week_monday_local.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start_utc = week_monday_local.astimezone(ZoneInfo("UTC"))
    week_end_utc = (week_monday_local + timedelta(days=7)).astimezone(ZoneInfo("UTC"))

    def _series_active_in_week(item, ws, we):
        t0 = getattr(item, "type", "")
        rec = getattr(item, "recurrence", None)
        rrule_string = getattr(rec, "rrule_string", None) if rec else None
        if not rrule_string:
            return False
        if is_birthday(item):
            return False
        if t0 in ("appointment", "event"):
            s0 = getattr(item, "start_utc", None)
            e0 = getattr(item, "end_utc", None)
            if s0:
                if e0 and (s0 < we) and (e0 > ws):
                    return True
                try:
                    from utils.rrule_helpers import enumerate_occurrences
                    occs = list(enumerate_occurrences(s0, e0, rrule_string, window_start=ws, window_end=we))
                    if occs:
                        return True
                except Exception:
                    if ws <= s0 < we:
                        return True
            return False
        if t0 in ("task", "reminder"):
            base_ts = getattr(item, "due_utc", None) or getattr(item, "reminder_utc", None)
            try:
                from utils.rrule_helpers import enumerate_occurrences
                occs = list(enumerate_occurrences(base_ts, None, rrule_string, window_start=ws, window_end=we)) if base_ts else []
                return bool(occs)
            except Exception:
                return bool(base_ts and (ws <= base_ts < we))
        return False

    for it in items:
        t = getattr(it, "type", "")
        has_recur = bool(getattr(it, "recurrence", None) and getattr(it.recurrence, "rrule_string", None))
        if has_recur and _series_active_in_week(it, week_start_utc, week_end_utc):
            recurring_items.append(it)

        # Panels Heute/Nächste 7/Überfällig
        if t == "appointment":
            s = getattr(it, "start_utc", None)
            e = getattr(it, "end_utc", None)
            if not s and not e:
                without_date.append(it)
            else:
                if s and (s <= win_end_48h):
                    upcoming.append(it)
                sL = as_local(s) if s else None
                if sL and (start_today <= sL <= end_today):
                    upcoming_today.append(it)
                elif sL and (start_today < sL <= end_7d):
                    upcoming_next7.append(it)
        elif t in ("task", "reminder"):
            ts = getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None)
            if not ts:
                without_date.append(it)
            else:
                # Überfällig-Check: vergangen UND nicht-terminal
                now_utc_ts = now_utc()
                is_overdue = (ts < now_utc_ts) and not status_svc.is_terminal(getattr(it, "status", ""))
                
                if is_overdue:
                    overdue.append(it)

                if ts <= win_end_48h:
                    upcoming.append(it)
                tsL = as_local(ts)
                if start_today <= tsL <= end_today:
                    upcoming_today.append(it)
                elif start_today < tsL <= end_7d:
                    upcoming_next7.append(it)

    # Ohne Termin: terminale ausblenden
    without_date = [it for it in without_date if not status_svc.is_terminal(it.status)]

    # Feiertage
    holidays_upcoming = get_next_holidays_de_ni(today, count=30)

    def _aware(dt):
        return dt if dt and dt.tzinfo else (dt.replace(tzinfo=timezone.utc) if dt else datetime.max.replace(tzinfo=timezone.utc))

    def sort_key_time(it):
        if getattr(it, "type", "") in ("appointment","event"):
            s, e = next_or_display_occurrence(it, now=datetime.now(timezone.utc))
            return _aware(s or e or datetime.max)
        else:
            return _aware(getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None) or datetime.max)

    def _overdue_key(t):
        prio = getattr(t, "priority", -1)
        due = getattr(t, "due_utc", None)
        created = getattr(t, "created_utc", datetime.max.replace(tzinfo=timezone.utc))
        group = 0 if due is not None else 1
        return (group, -int(prio if prio is not None else -1), due or datetime.max.replace(tzinfo=timezone.utc), created)

    # Sichtzeiten für Event-Panels
    _now = datetime.now(timezone.utc)
    def _with_disp_times(it, now):
        if getattr(it, "type", "") == "event":
            s, e = next_or_display_occurrence(it, now=now)
            if s or e:
                it2 = it.__class__(**it.__dict__)
                if s: it2.start_utc = s
                if e: it2.end_utc = e
                return it2
        return it

    upcoming = [_with_disp_times(x, _now) for x in upcoming]
    upcoming_today = [_with_disp_times(x, _now) for x in upcoming_today]
    upcoming_next7 = [_with_disp_times(x, _now) for x in upcoming_next7]
    overdue = [_with_disp_times(x, _now) for x in overdue]

    upcoming.sort(key=sort_key_time)
    overdue.sort(key=_overdue_key)
    upcoming_today.sort(key=sort_key_time)
    upcoming_next7.sort(key=sort_key_time)
    without_date.sort(key=lambda x: (
        -(x.priority or 0),  # Höchste Priorität zuerst (negativ: -5, -4, -3, -2, -1, 0)
        (x.last_modified_utc or datetime.min.replace(tzinfo=timezone.utc))  # Älteste Änderung zuerst (aufsteigend)
    ))


    def sort_key_recurring(it):
        t = getattr(it, "type", "")
        if t in ("appointment", "event"):
            s, e = next_or_display_occurrence(it, now=datetime.now(timezone.utc))
            return s or datetime.max.replace(tzinfo=timezone.utc)
        else:
            return getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None) or datetime.max.replace(tzinfo=timezone.utc)
    recurring_items.sort(key=sort_key_recurring)

    # Wochenkalender-Raster
    monday_this_week = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    monday_start = monday_this_week + timedelta(weeks=cal_week_offset)
    total_days = max(1, int(cal_weeks)) * 7
    d0_local = monday_start
    dN_local = d0_local + timedelta(days=total_days)
    d0_utc = d0_local.astimezone(ZoneInfo("UTC"))
    dN_utc = dN_local.astimezone(ZoneInfo("UTC"))

    # Occurrence-Expansion mit RRULE, Mehrtages-Spannen und Geburtstagen
    def expand_occurrences(it, start_utc, end_utc):
        out = []
        t = getattr(it, "type", "")
        if t in ("task","reminder"):
            ts = getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None)
            if ts and start_utc <= ts < end_utc:
                out.append((ts, None, it))
            return out

        s0 = getattr(it, "start_utc", None)
        e0 = getattr(it, "end_utc", None)
        recur = getattr(it, "recurrence", None)
        rrule_string = getattr(recur, "rrule_string", None) if recur else None

        # Geburtstage jährlich projizieren (auch ohne RRULE)
        is_bday = (t == "event") and ("geburtstag" in (getattr(it, "tags", []) or []))
        if is_bday and s0:
            s0_local = s0.astimezone(berlin)
            y_start = d0_utc.astimezone(berlin).year
            y_end = dN_utc.astimezone(berlin).year
            for y in range(y_start, y_end + 1):
                try:
                    occ_local = s0_local.replace(year=y)
                except ValueError:
                    if s0_local.month == 2 and s0_local.day == 29:
                        occ_local = s0_local.replace(year=y, day=28)
                    else:
                        continue
                occ_utc = occ_local.astimezone(ZoneInfo("UTC"))
                if start_utc <= occ_utc < end_utc:
                    occ_end_utc = None
                    if e0:
                        dur = e0 - s0
                        occ_end_utc = occ_utc + dur
                    out.append((occ_utc, occ_end_utc, it))
            return out

        if not s0 and not rrule_string:
            return out

        if not rrule_string:
            if s0 and (start_utc <= s0 < end_utc):
                out.append((s0, e0, it))
            return out

        try:
            from utils.rrule_helpers import enumerate_occurrences
            for occ_s, occ_e in enumerate_occurrences(s0, e0, rrule_string, window_start=start_utc, window_end=end_utc):
                out.append((occ_s, occ_e, it))
        except Exception:
            if s0 and (start_utc <= s0 < end_utc):
                out.append((s0, e0, it))
        return out

    # Tagesbuckets in lokaler TZ füllen (jede Occurrence einem Tag zuweisen)
    all_day_buckets: dict[date, list] = {}
    for it in items:
        occs = expand_occurrences(it, d0_utc, dN_utc)
        for occ_s, occ_e, src in occs:
            data = src.__dict__.copy()
            if getattr(src, "type", "") in ("appointment","event"):
                data["start_utc"] = occ_s
                data["end_utc"] = occ_e
            inst = src.__class__(**data)

            # Bei Mehrtages-Span: jeden betroffenen Tag befüllen
            if inst.type in ("appointment","event"):
                start_l = inst.start_utc.astimezone(berlin) if getattr(inst, "start_utc", None) else None
                end_l = inst.end_utc.astimezone(berlin) if getattr(inst, "end_utc", None) else start_l
                if not start_l:
                    continue
                d = start_l.date()
                last = end_l.date() if end_l else d
                while d <= last:
                    if d0_local.date() <= d < dN_local.date():
                        all_day_buckets.setdefault(d, []).append(inst)
                    d = d + timedelta(days=1)
            else:
                # task/reminder: Einzeltermin
                basis = getattr(inst, "due_utc", None) or getattr(inst, "reminder_utc", None)
                if not basis:
                    continue
                basis_local = basis.astimezone(berlin)
                key_date = basis_local.date()
                if d0_local.date() <= key_date < dN_local.date():
                    all_day_buckets.setdefault(key_date, []).append(inst)

    # Wochen/Tag-Grid
    headers = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    weeks = []
    for week_idx in range(max(1, int(cal_weeks))):
        week_start_local = monday_start + timedelta(weeks=week_idx)
        calendar_week_number = week_start_local.isocalendar()[1]
        week_days = []
        for day_num in range(7):
            day_local = week_start_local + timedelta(days=day_num)
            day_items = list(all_day_buckets.get(day_local.date(), []))
            def per_day_key(x):
                if x.type in ("appointment","event"):
                    return getattr(x, "start_utc", None) or getattr(x, "end_utc", None) or datetime.max.replace(tzinfo=timezone.utc)
                else:
                    return getattr(x, "due_utc", None) or getattr(x, "reminder_utc", None) or datetime.max.replace(tzinfo=timezone.utc)
            day_items.sort(key=per_day_key)
            week_days.append({
                "label": day_local.strftime("%d.%m."),
                "items": day_items,
                "is_weekend": day_local.weekday() in (5, 6),
                "is_today": (day_local.date() == today.date()),
                "date": day_local
            })
        weeks.append({"week_number": calendar_week_number, "days": week_days})
    calendar_week = {"headers": headers, "weeks": weeks, "offset": cal_week_offset, "count": max(1, int(cal_weeks))}

    # Status-Farben
    _raw = {t: status_svc.get_options_for(t) for t in ("task","reminder","appointment","event")}
    type_status_colors = {
        t: {sd.key: sd.color_light for sd in defs if getattr(sd, "color_light", None)}
        for t, defs in _raw.items()
    }

    # Nächste Ereignisse: Holidays + echte Events (2 Monate)
    horizon_2m_end = now_utc() + timedelta(days=60)
    def _start_key(ev):
        dt = getattr(ev, "start_utc", None) if hasattr(ev, "start_utc") else ev.get("start_utc")
        return dt or datetime.max.replace(tzinfo=timezone.utc)

    holiday_events = []
    holidays_upcoming.sort(key=_start_key)
    for ev in holidays_upcoming:
        s = ev.get("start_utc"); e = ev.get("end_utc")
        if s and (s < horizon_2m_end):
            holiday_events.append({
                "type": "event",
                "name": ev.get("name"),
                "start_utc": s,
                "end_utc": e,
                "priority": ev.get("priority", 0),
                "tags": list(ev.get("tags", [])),
                "status": ev.get("status", "EVENT_SCHEDULED"),
            })

    real_events = []
    for it in items:
        if getattr(it, "type", "") != "event":
            continue
        s, e = next_or_display_occurrence(it, now=now_utc())
        if not s and not e:
            continue
        is_running = (s and e and (s <= now_utc() < e))
        is_upcoming = (s and (now_utc() <= s <= horizon_2m_end))
        if is_running or is_upcoming:
            data = it.__dict__.copy()
            data["start_utc"] = s; data["end_utc"] = e
            it2 = it.__class__(**data)
            real_events.append(it2)

    def _start_key_mixed(ev):
        if hasattr(ev, "start_utc"):
            return getattr(ev, "start_utc") or _aware(datetime.max)
        return ev.get("start_utc") or _aware(datetime.max)

    events_next2m = sorted([*real_events, *holiday_events], key=_start_key_mixed)

    # Export-Hilfsfunktion (nutzt dieselbe Expansion)
    def build_calendar_rows(start_utc, end_utc):
        rows = []
        lok = berlin
        for it in items:
            for occ_s, occ_e, src in expand_occurrences(it, start_utc, end_utc):
                start_local = occ_s.astimezone(lok) if occ_s else None
                end_local = occ_e.astimezone(lok) if occ_e else None
                rows.append({
                    "Typ": getattr(src, "type", ""),
                    "Titel": getattr(src, "name", ""),
                    "Start (lokal)": start_local.strftime("%d.%m.%Y %H:%M") if start_local else "",
                    "Ende (lokal)": end_local.strftime("%d.%m.%Y %H:%M") if end_local else "",
                    "Start UTC": (occ_s.isoformat().replace("+00:00","Z") if occ_s else ""),
                    "Ende UTC": (occ_e.isoformat().replace("+00:00","Z") if occ_e else ""),
                    "Ganztägig": "ja" if bool(getattr(src, "is_all_day", False)) else "nein",
                    "Status": getattr(src, "status", "") or "",
                    "Priorität": getattr(src, "priority", None),
                    "Tags": ", ".join(getattr(src, "tags", []) or []),
                    "Links": ", ".join(getattr(src, "links", []) or []),
                    "Item-ID": getattr(src, "id", "") or "",
                })
        return rows

    header_today = format_local_weekday_de(today) + ", " + today.strftime("%d.%m.%Y")

    # Kontext
    ctx = {
        "request": request,
        "today": today,
        "now_local": today,
        "by_type": by_type,
        "by_status": by_status,
        "top_tags": Counter(t for it in items for t in (getattr(it, "tags", []) or []) if t).most_common(10),
        "overdue": overdue[:30],
        "upcoming_today": upcoming_today[:30],
        "upcoming_next7": upcoming_next7[:30],
        "without_date": without_date[:50],
        "recurring_items": recurring_items[:50],
        "events_next2m": events_next2m[:50],
        "calendar_week": {"headers": headers, "weeks": weeks, "offset": cal_week_offset, "count": max(1, int(cal_weeks))},
        "cal_week_offset": cal_week_offset,
        "cal_weeks": max(1, int(cal_weeks)),
        "status_display": lambda k: (status_svc.get_display_name(k) or (k or "")),
        "header_today": header_today,
        "type_status_colors": type_status_colors,
        "format_dashboard_time": format_dashboard_time,
        "is_birthday": is_birthday,
        "get_priority_class": get_priority_class,
        "is_overdue_item": is_overdue_item,
        "berlin_tz": berlin,
        "is_terminal_status": lambda k: status_svc.is_terminal(k),
    }

    return templates.TemplateResponse("dashboard.html", ctx)


@app.post("/tools/normalize_birthdays")
def normalize_birthdays(request: Request, confirm: int = 0, repo: DbRepository = Depends(get_repo)):
    items = repo.list_all()
    affected = []
    for it in items:
        if getattr(it, "type", "") != "event":
            continue
        if is_birthday(it):
            continue
        changes = {}
        if getattr(it, "is_all_day", None) is False:
            changes["set_is_all_day"] = True
        # Anzeige-Rollforward nur als Vorschau
        y = compute_next_yearly_from(it, now=datetime.now(timezone.utc))
        if y:
            s, e = y
            changes["next_display_start"] = s.isoformat()
            changes["next_display_end"] = e.isoformat()
        if changes:
            affected.append({"id": it.id, "name": it.name, "changes": changes})
    if not confirm:
        return {"preview": True, "count": len(affected), "items": affected}
    # Persistiere nur is_all_day=True
    updated = 0
    for rec in affected:
        if rec["changes"].get("set_is_all_day"):
            it = repo.get(rec["id"])
            it2 = it.__class__(**{**it.__dict__, "is_all_day": True})
            repo.upsert(it2)
            updated += 1
    repo.conn.commit()
    return {"applied": True, "updated": updated, "skipped": len(affected) - updated}

@app.get("/dashboard/export.xlsx")
def export_dashboard_excel(
    request: Request,
    q: str | None = None,
    types: str | None = None,
    status_keys: str | None = None,
    status: str | None = None,
    show_private: str | int | None = None,
    tags: str | None = None,
    cal_weeks: str | int | None = None,
    cal_week_offset: str | int | None = None,
    repo: DbRepository = Depends(get_repo),
):
    # Erfordert openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    berlin = ZoneInfo("Europe/Berlin")

    # Daten + Privatsphäre wie im Dashboard
    items = repo.list_all()
    include_private = bool(int(show_private or 0))
    def _visible_by_privacy(it) -> bool:
        is_private = bool(getattr(it, "private", False))
        return include_private or not is_private
    items = [it for it in items if _visible_by_privacy(it)]

    # Wochenfenster wie im Dashboard
    base_local = now_utc().astimezone(berlin)
    monday_this_week = (base_local - timedelta(days=base_local.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    offset = int(cal_week_offset or 0)
    weeks = max(1, int(cal_weeks or 1))
    monday_start = monday_this_week + timedelta(weeks=offset)

    d0_local = monday_start
    dN_local = d0_local + timedelta(days=weeks * 7)
    d0_utc = d0_local.astimezone(ZoneInfo("UTC"))
    dN_utc = dN_local.astimezone(ZoneInfo("UTC"))

    # Occurrence-Expansion wie Dashboard
    def expand_occurrences(it, start_utc, end_utc):
        out = []
        t = getattr(it, "type", "")

        if t in ("task","reminder"):
            ts = getattr(it, "due_utc", None) or getattr(it, "reminder_utc", None)
            if ts and start_utc <= ts < end_utc:
                out.append((ts, None, it))
            return out

        s0 = getattr(it, "start_utc", None)
        e0 = getattr(it, "end_utc", None)
        recur = getattr(it, "recurrence", None)
        rrule_string = getattr(recur, "rrule_string", None) if recur else None

        if not s0 and not rrule_string:
            return out

        if not rrule_string:
            if s0 and (start_utc <= s0 < end_utc):
                out.append((s0, e0, it))
            return out

        try:
            from utils.rrule_helpers import enumerate_occurrences
            for occ_s, occ_e in enumerate_occurrences(s0, e0, rrule_string, window_start=start_utc, window_end=end_utc):
                out.append((occ_s, occ_e, it))
        except Exception:
            if s0 and (start_utc <= s0 < end_utc):
                out.append((s0, e0, it))
        return out

    # Buckets (lokal) pro Tag
    all_day_buckets = {}  # key = date() -> [inst,...]
    for it in items:
        for occ_s, occ_e, src in expand_occurrences(it, d0_utc, dN_utc):
            data = src.__dict__.copy()
            if getattr(src, "type", "") in ("appointment","event"):
                data["start_utc"] = occ_s
                data["end_utc"] = occ_e
            inst = src.__class__(**data)
            if inst.type in ("appointment","event"):
                basis = inst.start_utc or inst.end_utc
            else:
                basis = getattr(inst, "due_utc", None) or getattr(inst, "reminder_utc", None)
            if not basis:
                continue
            key_date = basis.astimezone(berlin).date()
            all_day_buckets.setdefault(key_date, []).append(inst)

    # Sortierschlüssel pro Tag wie im Dashboard
    def per_day_key(x):
        if x.type in ("appointment","event"):
            return getattr(x, "start_utc", None) or getattr(x, "end_utc", None) or datetime.max.replace(tzinfo=timezone.utc)
        else:
            return getattr(x, "due_utc", None) or getattr(x, "reminder_utc", None) or datetime.max.replace(tzinfo=timezone.utc)

    # Anzeige-Labels
    def status_label(k: str | None) -> str:
        return status_svc.get_display_name(k) or (k or "")
    def prio_label(p: int | None) -> str:
        mapping = {
            0: "Keine",
            1: "Niedrig",
            2: "Normal",
            3: "Hoch",
            4: "Kritisch",
            5: "Blockierend",
        }
        return f"Prio: {mapping.get(p, '—')}"

    # Excel Setup
    wb = Workbook()
    ws = wb.active
    ws.title = f"KW {d0_local.isocalendar()[1]}"

    headers = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    widths = [26, 26, 26, 26, 26, 26, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wrap_top = Alignment(vertical="top", horizontal="left", wrap_text=True)
    center = Alignment(vertical="center", horizontal="center")
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))
    bold = Font(bold=True)

    row_cursor = 1
    for week_idx in range(weeks):
        week_start_local = d0_local + timedelta(days=7*week_idx)
        kw = week_start_local.isocalendar()[1]

        # Wochenkopf
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=7)
        c = ws.cell(row=row_cursor, column=1, value=f"Kalenderwoche {kw}")
        c.font = bold
        c.alignment = center
        row_cursor += 1

        # Tageskopf
        for col, h in enumerate(headers, start=1):
            day_local = week_start_local + timedelta(days=col-1)
            c = ws.cell(row=row_cursor, column=col, value=f"{h} {day_local.strftime('%d.%m.')}")
            c.font = bold
            c.fill = header_fill
            c.alignment = center
            c.border = border
        row_cursor += 1

        # Zeilenweise Items: eine Zeile pro Item, in der passenden Tages-Spalte
        # Baue pro Tag die gerenderten Zeilen (Strings)
        per_day_lines = []
        max_rows = 0
        for col in range(7):
            day_local = week_start_local + timedelta(days=col)
            day_items = list(all_day_buckets.get(day_local.date(), []))
            day_items.sort(key=per_day_key)
            lines = []
            for inst in day_items:
                t = inst.type
                name = inst.name or ""
                # Zeitdarstellung
                if t in ("appointment","event"):
                    start = getattr(inst, "start_utc", None)
                    end = getattr(inst, "end_utc", None)
                    sl = start.astimezone(berlin).strftime("%H:%M") if start else "--:--"
                    el = end.astimezone(berlin).strftime("%H:%M") if end else ""
                    when = f"{sl}-{el}".strip("-")
                else:
                    due = getattr(inst, "due_utc", None) or getattr(inst, "reminder_utc", None)
                    when = due.astimezone(berlin).strftime("%H:%M") if due else ""
                # Labels
                # Labels
                pr = prio_label(getattr(inst, "priority", None))             # z. B. "Prio: Normal"
                st = status_label(getattr(inst, "status", None))              # z. B. "Geplant" oder "Erledigt"

                # Kopf: "12:00 Uhr (Geplant) · Prio: Normal"
                # Uhrzeitformat "HH:MM Uhr"; für Spannen wie "08:30-09:15" kannst du bei Terminen auch beide Zeiten formatieren
                def fmt_time_label(tstr: str) -> str:
                    return (tstr + " Uhr") if tstr and ":" in tstr and "-" not in tstr else tstr

                head = " ".join(x for x in [fmt_time_label(when), f"({st})" if st else "", f"{pr}" if pr else ""] if x).strip()

                # Zweite Zeile: der Titel
                title = (inst.name or "").strip()

                # Zusammenführen mit Zeilenumbruch
                line = head + ("\n" + title if title else "")
                lines.append(line)

            per_day_lines.append(lines)
            max_rows = max(max_rows, len(lines))

        # Für die maximale Anzahl Items Zeilen anlegen
        for i in range(max_rows):
            for col in range(1, 8):
                text = per_day_lines[col-1][i] if i < len(per_day_lines[col-1]) else ""
                c = ws.cell(row=row_cursor, column=col, value=text)
                c.alignment = wrap_top
                c.border = border
            row_cursor += 1

        # kleine Lücke zwischen Wochen
        row_cursor += 1

    ws.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f'kalender_kw{d0_local.isocalendar()[1]}_w{weeks}.xlsx'
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "Cache-Control": "no-store"}
    )

@app.get("/items/{item_id}/zoom", response_class=HTMLResponse)
def zoom_item_description(item_id: str, request: Request, repo: DbRepository = Depends(get_repo)):
    it = repo.get(item_id)
    if not it:
        raise HTTPException(404, "Item nicht gefunden")

    # Name/Titel
    title = f"Details – {getattr(it, 'name', '') or 'Unbenannt'}"

    # Beschreibung als HTML. Falls du Markdown nutzt, hier rendern; sonst escapen.
    from markupsafe import escape
    desc_raw = getattr(it, "description", "") or ""
    # Optional: Markdown-Konvertierung, falls vorhanden:
    # html_body = markdown(desc_raw)  # nur wenn markdown() vorhanden
    html_body = f"<pre style='white-space:pre-wrap; font:inherit; margin:0;'>{escape(desc_raw)}</pre>"

    html = f"""<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <title>{escape(title)}</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body {{ margin: 1rem; font: 14.5px/1.5 system-ui, Arial, sans-serif; color:#111; background:#fff; }}
    h1 {{ font-size: 1.1rem; margin:0 0 .75rem; color:#666; }}
    .content {{ white-space: normal; }}
  </style>
</head>
<body>
  <h1>{escape(title)}</h1>
  <div class="content">{html_body}</div>
</body>
</html>"""
    return HTMLResponse(content=html, status_code=200)
